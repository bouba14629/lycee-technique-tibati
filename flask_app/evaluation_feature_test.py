import os
import subprocess
import tempfile
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

os.environ["DATABASE_URL"] = "sqlite:////tmp/ltt-evaluation-feature.sqlite"
os.environ["LTT_ENV"] = "development"
os.environ["LTT_INITIAL_ADMIN_PASSWORD"] = "FoundateurTest#2026"

from app import app
from models import Course, Department, Grade, PlannedAssessment, SchoolClass, Section, Student, Subject, Teacher, User, db
from utils import dashboard_alerts


def add_user(username, full_name, role):
    user = User(username=username, full_name=full_name, role=role, active=True)
    user.set_password("Test#2026")
    db.session.add(user)
    db.session.flush()
    return user


def switch_user(client, user):
    user.session_token = f"test-session-{user.id}"
    db.session.commit()
    with client.session_transaction() as session:
        session.clear()
        session["user_id"] = user.id
        session["role"] = user.role
        session["session_token"] = user.session_token


with app.app_context():
    db.drop_all()
    db.create_all()
    director = add_user("proviseur.eval", "Proviseur Évaluation", "directeur")
    director_id = director.id
    censeur = add_user("censeur.eval", "Censeur Évaluation", "censeur")
    censeur_id = censeur.id
    teacher_user = add_user("enseignant.eval", "Enseignant Évaluation", "enseignant")
    student_user = add_user("eleve.eval", "Élève Évaluation", "eleve")
    student_user_id = student_user.id
    section = Section(name="Section Évaluation", code="EVAL")
    db.session.add(section)
    db.session.flush()
    department = Department(name="Département Évaluation", code="DEP", section_id=section.id)
    db.session.add(department)
    db.session.flush()
    censeur.section_id = section.id
    other_department = Department(name="Département Écarté", code="EC", section_id=section.id)
    db.session.add(other_department)
    db.session.flush()
    school_class = SchoolClass(name="Classe Évaluation", code="EVAL-CLASS", level="Test", department_id=department.id)
    other_class = SchoolClass(name="Classe Écartée", code="ECART-CLASS", level="Test", department_id=other_department.id)
    subject = Subject(name="Mathématiques appliquées", department_id=department.id)
    other_subject = Subject(name="Maintenance", department_id=other_department.id)
    teacher = Teacher(user_id=teacher_user.id, department_id=department.id)
    db.session.add_all([school_class, other_class, subject, other_subject, teacher])
    db.session.flush()
    course = Course(subject_id=subject.id, teacher_id=teacher.id, class_id=school_class.id)
    other_course = Course(subject_id=other_subject.id, teacher_id=teacher.id, class_id=other_class.id)
    pending_subject = Subject(name="Dessin technique", department_id=department.id)
    db.session.add(pending_subject)
    db.session.flush()
    pending_course = Course(subject_id=pending_subject.id, teacher_id=teacher.id, class_id=school_class.id)
    student = Student(user_id=student_user.id, class_id=school_class.id, matricule="EVAL-001", first_name="Élève", last_name="Évaluation")
    db.session.add_all([course, other_course, pending_course, student])
    db.session.commit()

    with app.test_client() as client:
        switch_user(client, censeur)
        plan = client.post("/pedagogie/evaluations", data={
            "term": "Trimestre 1", "course_id": course.id, "sequence": 1,
            "title": "Évaluation diagnostique", "scheduled_date": date.today().isoformat(), "max_value": "20",
        }, follow_redirects=False)
        assert plan.status_code == 302
        assessment = PlannedAssessment.query.one()
        assert assessment.status == "Planifiée" and assessment.sequence == 1
        invalid_plan = client.post("/pedagogie/evaluations", data={
            "term": "Trimestre 1", "course_id": course.id, "sequence": 2,
            "title": "Évaluation hors barème", "scheduled_date": date.today().isoformat(), "max_value": "21",
        }, follow_redirects=True)
        assert invalid_plan.status_code == 200
        assert b"bar\xc3\xa8me d\xe2\x80\x99une \xc3\xa9valuation doit \xc3\xaatre compris entre 1 et 20" in invalid_plan.data
        assert PlannedAssessment.query.count() == 1
        assert any(alert[2] == "Évaluations à finaliser" for alert in dashboard_alerts())
        assert client.get("/pedagogie/evaluations?term=Trimestre+1").status_code == 200

        switch_user(client, teacher_user)
        assessment_list = client.get("/enseignant/evaluations?term=Trimestre+1")
        assert assessment_list.status_code == 200 and b"diagnostique" in assessment_list.data
        grade_form = client.get(f"/enseignant/notes/{course.id}?term=Trimestre+1")
        assert grade_form.status_code == 200 and b"valuation planifi" in grade_form.data
        submit = client.post(f"/enseignant/notes/{course.id}", data={
            "term": "Trimestre 1", "assessment_id": assessment.id, f"grade_{student.id}": "15",
        }, follow_redirects=False)
        assert submit.status_code == 302
        grade = Grade.query.one()
        assert grade.sequence == 1 and grade.value == 15
        rejected_grade = client.post(f"/enseignant/notes/{course.id}", data={
            "term": "Trimestre 1", "assessment_id": assessment.id, f"grade_{student.id}": "21",
        }, follow_redirects=True)
        assert rejected_grade.status_code == 200
        assert b"note doit \xc3\xaatre comprise entre 0 et 20" in rejected_grade.data
        assert Grade.query.one().value == 15
        assert assessment.status == "Saisie complète" and assessment.submitted_at is not None
        assert not any(alert[2] == "Évaluations à finaliser" for alert in dashboard_alerts())

        switch_user(client, student_user)
        student_bulletin = client.get("/eleve/bulletin?term=Trimestre+1")
        assert student_bulletin.status_code == 200 and b"Bulletin priv" in student_bulletin.data

        switch_user(client, director)
        tracking = client.get("/pedagogie/evaluations?term=Trimestre+1")
        assert tracking.status_code == 200 and b"Saisie compl" in tracking.data
        assert b"Synth\xc3\xa8se par classe" in tracking.data and b"100%" in tracking.data
        indicators = client.get(f"/censeur/indicateurs?term=Trimestre+1&sequence=1&assessment_id={assessment.id}")
        assert indicators.status_code == 200
        assert b"R\xc3\xa9sultats des \xc3\xa9valuations" in indicators.data and b"100%" in indicators.data
        sequence_stats = client.get("/censeur/conseil-de-classe/statistiques?term=Trimestre+1&sequence=1")
        assert sequence_stats.status_code == 200 and b"S\xc3\xa9quence 1" in sequence_stats.data
        sequence_export = client.get("/censeur/conseil-de-classe/statistiques/export.xlsx?term=Trimestre+1&sequence=1")
        assert sequence_export.status_code == 200 and sequence_export.data[:2] == b"PK"
        detailed_stats_url = f"/censeur/conseil-de-classe/statistiques?term=Trimestre+1&class_id={school_class.id}&course_id={course.id}&assessment_id={assessment.id}"
        detailed_stats = client.get(detailed_stats_url)
        assert detailed_stats.status_code == 200 and b"15" in detailed_stats.data
        assert b'name="assessment_id"' not in detailed_stats.data
        assert b"Toutes les \xc3\xa9valuations" not in detailed_stats.data
        assert b"Fiche centr\xc3\xa9e sur" not in detailed_stats.data
        detailed_export = client.get(f"/censeur/conseil-de-classe/statistiques/export.xlsx?term=Trimestre+1&class_id={school_class.id}&course_id={course.id}&assessment_id={assessment.id}")
        assert detailed_export.status_code == 200 and detailed_export.data[:2] == b"PK"
        db.session.remove()
        switch_user(client, db.session.get(User, director_id))
        director_refresh = client.get("/pedagogie/evaluations?term=Trimestre+1&sequence=1")
        assert director_refresh.status_code == 200 and b"Saisie compl" in director_refresh.data and b"100%" in director_refresh.data
        assert b"Synth\xc3\xa8se par classe" in director_refresh.data
        assert b"1/1" in director_refresh.data and b"<td>0</td>" in director_refresh.data
        switch_user(client, db.session.get(User, censeur_id))
        refreshed_stats = client.get(detailed_stats_url)
        assert refreshed_stats.status_code == 200 and b"15" in refreshed_stats.data
        department_stats = client.get(
            f"/censeur/conseil-de-classe/statistiques?term=Trimestre+1&department_id={department.id}"
        )
        assert department_stats.status_code == 200
        assert b"EVAL-CLASS" in department_stats.data
        assert b"ECART-CLASS" not in department_stats.data
        class_subject_stats = client.get(
            f"/censeur/conseil-de-classe/statistiques?term=Trimestre+1&class_ids={school_class.id}"
        )
        assert class_subject_stats.status_code == 200
        assert "Mathématiques appliquées".encode() in class_subject_stats.data
        assert "Maintenance".encode() not in class_subject_stats.data
        assert f'data-class-ids="{school_class.id}"'.encode() in class_subject_stats.data
        assert b"filterCouncilSubjects" in class_subject_stats.data
        unrelated_subject = client.get(
            f"/censeur/conseil-de-classe/statistiques?term=Trimestre+1&class_ids={school_class.id}&subject_ids={other_subject.id}"
        )
        assert unrelated_subject.status_code == 403
        department_export = client.get(
            f"/censeur/conseil-de-classe/statistiques/export.xlsx?term=Trimestre+1&department_id={department.id}"
        )
        assert department_export.status_code == 200 and department_export.data[:2] == b"PK"
        department_indicators = client.get(
            f"/censeur/indicateurs?term=Trimestre+1&department_id={department.id}"
        )
        assert department_indicators.status_code == 200
        assert b"EVAL-CLASS" in department_indicators.data
        assert b"ECART-CLASS" not in department_indicators.data
        department_indicators_export = client.get(
            f"/censeur/indicateurs/export.xlsx?term=Trimestre+1&department_id={department.id}"
        )
        assert department_indicators_export.status_code == 200 and department_indicators_export.data[:2] == b"PK"
        multiple_filters = (
            f"class_ids={school_class.id}&class_ids={other_class.id}"
            f"&subject_ids={subject.id}&subject_ids={other_subject.id}"
        )
        multiple_stats = client.get(f"/censeur/conseil-de-classe/statistiques?term=Trimestre+1&{multiple_filters}")
        assert multiple_stats.status_code == 200
        assert b"EVAL-CLASS" in multiple_stats.data and b"ECART-CLASS" in multiple_stats.data
        assert b'name="subject_ids"' in multiple_stats.data
        assert "Mathématiques appliquées".encode() in multiple_stats.data and "Maintenance".encode() in multiple_stats.data
        multiple_stats_export = client.get(f"/censeur/conseil-de-classe/statistiques/export.xlsx?term=Trimestre+1&{multiple_filters}")
        assert multiple_stats_export.status_code == 200 and multiple_stats_export.data[:2] == b"PK"
        multiple_indicators = client.get(f"/censeur/indicateurs?term=Trimestre+1&{multiple_filters}")
        assert multiple_indicators.status_code == 200
        assert b"EVAL-CLASS" in multiple_indicators.data and b"ECART-CLASS" in multiple_indicators.data
        assert b'name="subject_ids"' in multiple_indicators.data
        multiple_indicators_export = client.get(f"/censeur/indicateurs/export.xlsx?term=Trimestre+1&{multiple_filters}")
        assert multiple_indicators_export.status_code == 200 and multiple_indicators_export.data[:2] == b"PK"

        switch_user(client, db.session.get(User, censeur_id))
        xlsx_bulletin = client.get(f"/eleves/{student.id}/bulletin/telecharger.xlsx?term=Trimestre+1")
        assert xlsx_bulletin.status_code == 200 and xlsx_bulletin.data[:2] == b"PK"
        workbook = load_workbook(BytesIO(xlsx_bulletin.data), data_only=True)
        xlsx_values = " ".join(str(cell.value) for row in workbook.active.iter_rows() for cell in row if cell.value is not None)
        assert "15" in xlsx_values
        assert "Dessin technique" in xlsx_values and "Notes en attente" in xlsx_values
        pdf_bulletin = client.get(f"/eleves/{student.id}/bulletin/telecharger?term=Trimestre+1")
        assert pdf_bulletin.status_code == 200 and pdf_bulletin.data[:4] == b"%PDF"
        preview_path = os.environ.get("LTT_BULLETIN_PREVIEW_PATH")
        if preview_path:
            with open(preview_path, "wb") as preview_file:
                preview_file.write(pdf_bulletin.data)
        with tempfile.NamedTemporaryFile(suffix=".pdf") as pdf_file:
            pdf_file.write(pdf_bulletin.data)
            pdf_file.flush()
            pdf_text = subprocess.run(["pdftotext", pdf_file.name, "-"], check=True, capture_output=True, text=True).stdout
        assert "15" in pdf_text
        assert "Dessin technique" in pdf_text and "Notes en attente" in pdf_text

        switch_user(client, db.session.get(User, student_user_id))
        assert client.get("/pedagogie/evaluations").status_code == 403
        assert client.get("/enseignant/evaluations").status_code == 403

print("EVALUATION_FEATURE_TEST_OK")
