from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

NAVY = "0B2545"
GOLD = "C9A227"
CREAM = "F8F5EC"

thin = Side(style="thin", color="D9D0B0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _header(ws, row, cols, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def bulletin_workbook(student, data, term):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulletin"
    for i, w in enumerate([28, 12, 12, 12, 10, 12, 8, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _header(ws, 1, 8, "LYCÉE TECHNIQUE DE TIBATI")
    from utils import get_current_school_year
    ws.cell(row=2, column=1, value=f"Bulletin — {term} — Année scolaire {get_current_school_year()}").font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    r = 4
    meta = [
        ("Élève", student.full_name), ("Matricule", student.matricule),
        ("Classe", student.school_class.name if student.school_class else "—"),
        ("Redoublant", "Oui" if student.is_repeater else "Non"),
    ]
    for label, val in meta:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=9, color="6B7688")
        ws.cell(row=r, column=2, value=val).font = Font(size=10)
        r += 1
    r += 1

    headers = ["Matière", "Notes Trim.", f"Éval.{data.get('term_seq_a', '')}", f"Éval.{data.get('term_seq_b', '')}", "Coef.", "Notes×Coef.", "Rang", "Appréciation"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, size=9, color=NAVY)
        c.fill = PatternFill("solid", fgColor=CREAM)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    r += 1

    for cat in data["categories"]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cc = ws.cell(row=r, column=1, value=cat["name"].upper())
        cc.font = Font(bold=True, size=9.5, color=NAVY)
        cc.fill = PatternFill("solid", fgColor="EFE9D8")
        r += 1
        for row in cat["rows"]:
            vals = [row["course"].subject.name, row["notes_trim"] if row["notes_trim"] is not None else "En attente", row["eval_a"], row["eval_b"],
                    row["coef"], row["points"], row["rank"], row["appreciation"]]
            for i, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=i, value=v)
                c.border = BORDER
                c.font = Font(size=9.5)
            r += 1
        ws.cell(row=r, column=1, value=f"Total {cat['name']}").font = Font(bold=True, size=9, italic=True)
        ws.cell(row=r, column=5, value=cat["total_coef"]).font = Font(bold=True, size=9)
        ws.cell(row=r, column=6, value=cat["total_points"]).font = Font(bold=True, size=9)
        moyenne = cat["moyenne"] if cat["moyenne"] is not None else "—"
        attente = f" — {cat['pending_count']} matière(s) en attente" if cat.get("pending_count") else ""
        ws.cell(row=r, column=8, value=f"Moyenne : {moyenne}/20{attente}").font = Font(bold=True, size=9)
        r += 2

    ws.cell(row=r, column=1, value="MOYENNE GÉNÉRALE").font = Font(bold=True, size=12, color=NAVY)
    overall_avg_display = data['overall_avg'] if data['overall_avg'] is not None else '—'
    ws.cell(row=r, column=6, value=f"{overall_avg_display}/20").font = Font(bold=True, size=13, color=NAVY)
    r += 1
    rank_display = data['rank'] if data['rank'] is not None else '—'
    ws.cell(row=r, column=1, value=f"Rang : {rank_display} sur {data['class_size']}").font = Font(size=10)
    r += 2

    ws.cell(row=r, column=1, value="Absences justifiées (h)").font = Font(size=9, bold=True)
    ws.cell(row=r, column=2, value=data["absences_justified_hours"])
    ws.cell(row=r, column=3, value="Absences non justifiées (h)").font = Font(size=9, bold=True)
    ws.cell(row=r, column=4, value=data["absences_non_justified_hours"])
    r += 1
    ws.cell(row=r, column=1, value="Retards").font = Font(size=9, bold=True)
    ws.cell(row=r, column=2, value=data["retards"])

    wb_io = BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io


def users_workbook(users, title, role_filter=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"
    widths = [24, 18, 16, 26, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _header(ws, 1, len(widths), title)
    r = 3
    headers = ["Nom complet", "Identifiant", "Rôle", "Email", "Téléphone", "Statut"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, size=9.5, color=NAVY)
        c.fill = PatternFill("solid", fgColor=CREAM)
        c.border = BORDER
    r += 1
    for u in users:
        vals = [u.full_name, u.username, u.role_label, u.email or "", u.phone or "",
                "Actif" if u.active else "Désactivé"]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            c.font = Font(size=9.5)
        r += 1
    wb_io = BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io


def council_stats_workbook(stats, term):
    """Reproduit le modèle officiel 'FICHE STATISTIQUE DES RESULTATS DU TRIMESTRE' (Inscrits/Évalués/
    Moyenne≥10/%Réussite/Nbre TH/Encouragement/Félicitations/Moyenne générale/Forte-Faible moyenne/Observations)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fiche statistique"
    widths = [18] + [7] * 21 + [9] + [7] * 4 + [24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _header(ws, 1, len(widths), f"FICHE STATISTIQUE DES RÉSULTATS — {term}")

    r = 3
    groups = [("CODE", 1, 1), ("INSCRITS", 2, 4), ("ÉVALUÉS", 5, 7), ("MOYENNE ≥10", 8, 10),
              ("% RÉUSSITE", 11, 13), ("NBRE TH", 14, 16), ("ENCOURAGEMENT", 17, 19), ("FÉLICITATIONS", 20, 22),
              ("MOY. GÉN.", 23, 23), ("FORTE MOY.", 24, 25), ("FAIBLE MOY.", 26, 27), ("OBSERVATIONS", 28, 28)]
    for label, c1, c2 in groups:
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=label)
        cell.font = Font(bold=True, size=8.5, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=CREAM)
        cell.alignment = Alignment(horizontal="center")
    r += 1
    headers = [""] + ["F", "G", "T"] * 7 + [""] + ["F", "G"] * 2 + [""]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, size=8)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    r += 1

    def write_row(label, row, bold=False):
        nonlocal r
        vals = [label,
                row["inscrits"]["F"], row["inscrits"]["G"], row["inscrits"]["T"],
                row["evalues"]["F"], row["evalues"]["G"], row["evalues"]["T"],
                row["reussite"]["F"], row["reussite"]["G"], row["reussite"]["T"],
                row["pct"]["F"], row["pct"]["G"], row["pct"]["T"],
                row["th"]["F"], row["th"]["G"], row["th"]["T"],
                row["encouragement"]["F"], row["encouragement"]["G"], row["encouragement"]["T"],
                row["felicitations"]["F"], row["felicitations"]["G"], row["felicitations"]["T"],
                row["moyenne_generale"],
                row["forte_moyenne"]["F"], row["forte_moyenne"]["G"],
                row["faible_moyenne"]["F"], row["faible_moyenne"]["G"],
                row["observation"]]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            c.font = Font(bold=bold, size=8.5)
            if bold:
                c.fill = PatternFill("solid", fgColor=CREAM)
        r += 1

    if stats["cycle1"]:
        ws.cell(row=r, column=1, value="PREMIER CYCLE").font = Font(bold=True, size=9, color="FFFFFF")
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(widths))
        r += 1
        for row in stats["cycle1"]:
            write_row(row['class'].code or 'Sans code', row)
        write_row("TOTAL 1", stats["total_cycle1"], bold=True)
    if stats["cycle2"]:
        ws.cell(row=r, column=1, value="SECOND CYCLE").font = Font(bold=True, size=9, color="FFFFFF")
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(widths))
        r += 1
        for row in stats["cycle2"]:
            write_row(row['class'].code or 'Sans code', row)
        write_row("TOTAL 2", stats["total_cycle2"], bold=True)
    if stats["total_general"]:
        write_row("TOTAL GÉNÉRAL", stats["total_general"], bold=True)

    wb_io = BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io


def indicators_workbook(rows, totals, term, custom_types=None):
    """Reproduit fidèlement le modèle 'SYNTHÈSE - INDICATEURS PÉDAGOGIQUES' fourni par l'établissement :
    couverture des heures, des programmes (dont digital) et des travaux pratiques (dont digitalisés) —
    plus, le cas échéant, les indicateurs personnalisés créés par le Censeur, en colonnes supplémentaires."""
    custom_types = custom_types or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicateurs"
    widths = [24, 14, 20, 8, 8, 7, 9, 8, 7, 9, 8, 7, 8, 9, 7, 9, 9, 7] + [9, 9, 7] * len(custom_types)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _header(ws, 1, len(widths), f"SYNTHÈSE — INDICATEURS PÉDAGOGIQUES — {term}")

    r = 3
    # groupes de colonnes
    groups = [
        ("ENSEIGNANT", 1, 1),
        ("CODE / CLASSE", 2, 2),
        ("MATIÈRE", 3, 3),
        ("COUVERTURE DES HEURES", 4, 6),
        ("COUVERTURE DES PROGRAMMES", 7, 9),
        ("LEÇONS DIGITALISÉES", 10, 12),
        ("TRAVAUX PRATIQUES", 13, 15),
        ("TP DIGITALISÉS", 16, 18),
    ]
    col = 19
    for ct in custom_types:
        groups.append((ct.label.upper(), col, col + 2))
        col += 3
    for label, c1, c2 in groups:
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=label)
        cell.font = Font(bold=True, size=9.5, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=CREAM)
        cell.alignment = Alignment(horizontal="center")
    r += 1
    headers = ["Enseignant", "Code / Classe", "Matière",
               "Dues", "Faites", "%",
               "Prévues", "Faites", "%",
               "Prévues", "Faites", "%",
               "Prévus", "Réalisés", "%",
               "Prévus", "Réalisés", "%"]
    for ct in custom_types:
        headers += [ct.unit_planned, ct.unit_done, "%"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=CREAM)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    r += 1
    for row in rows:
        ind = row["ind"]
        vals = [row['teacher'].user.full_name, f"{row['course'].school_class.code or 'Sans code'} — {row['course'].school_class.name}", row['course'].subject.name,
                ind.hours_due, ind.hours_done, row["pct_hours"],
                ind.lessons_planned, ind.lessons_done, row["pct_lessons"],
                ind.digital_lessons_planned, ind.digital_lessons_done, row["pct_digital_lessons"],
                ind.tp_planned, ind.tp_done, row["pct_tp"],
                ind.digital_tp_planned, ind.digital_tp_done, row["pct_digital_tp"]]
        for ct in custom_types:
            cv = row.get("custom", {}).get(ct.id)
            vals += [cv.planned if cv else None, cv.done if cv else None, cv.pct if cv else None]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            c.font = Font(size=9)
        r += 1
    # ligne total
    total_vals = ["TOTAL", "", "",
                  totals["hours_due"], totals["hours_done"], totals["pct_hours"],
                  totals["lessons_planned"], totals["lessons_done"], totals["pct_lessons"],
                  totals["digital_lessons_planned"], totals["digital_lessons_done"], totals["pct_digital_lessons"],
                  totals["tp_planned"], totals["tp_done"], totals["pct_tp"],
                  totals["digital_tp_planned"], totals["digital_tp_done"], totals["pct_digital_tp"]]
    total_vals += ["", "", ""] * len(custom_types)
    for i, v in enumerate(total_vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(bold=True, size=9.5)
        c.fill = PatternFill("solid", fgColor=CREAM)
        c.border = BORDER

    wb_io = BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io


def sanctions_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sanctions"
    widths = [14, 24, 16, 10, 14, 30, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _header(ws, 1, len(widths), "LYCÉE TECHNIQUE DE TIBATI — ÉLÈVES SANCTIONNÉS")
    r = 3
    headers = ["Matricule", "Nom complet", "Classe", "Nb sanctions", "Heures d'absence", "Dernière sanction", "Date"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, size=9.5, color=NAVY)
        c.fill = PatternFill("solid", fgColor=CREAM)
        c.border = BORDER
    r += 1
    for row in rows:
        s = row["student"]
        vals = [s.matricule, s.full_name, s.school_class.name if s.school_class else "",
                row["nb_sanctions"], row["absence_hours"], row["last_sanction"].type,
                row["last_sanction"].date.strftime("%d/%m/%Y") if row["last_sanction"].date else ""]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            c.font = Font(size=9.5)
        r += 1
    wb_io = BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io


def absence_hours_workbook(school_class, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Absences"
    widths = [16, 28, 18, 18, 32]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    _header(ws, 1, len(widths), f"LYCÉE TECHNIQUE DE TIBATI — ABSENCES — {school_class.name}")
    headers = ["Matricule", "Nom complet", "Nombre d'absences", "Heures cumulées", "Dernières absences"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=CREAM)
        cell.border = BORDER
    for row_index, row in enumerate(rows, start=4):
        recent = ", ".join(record.date.strftime("%d/%m/%Y") for record in row["records"][:3])
        values = [row["student"].matricule, row["student"].full_name, row["count"], row["hours"], recent]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.border = BORDER
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def students_workbook(students, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Élèves"
    widths = [14, 24, 8, 14, 16, 18, 30, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _header(ws, 1, len(widths), title)
    r = 3
    headers = ["Matricule", "Nom complet", "Sexe", "Date naissance", "Classe", "Statut", "Adresse", "Parent(s)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, size=9.5, color=NAVY)
        c.fill = PatternFill("solid", fgColor=CREAM)
        c.border = BORDER
    r += 1
    for s in students:
        vals = [s.matricule, s.full_name, s.sex or "", s.dob.strftime("%d/%m/%Y") if s.dob else "",
                s.school_class.name if s.school_class else "", s.status, s.address or "",
                ", ".join(p.user.full_name for p in s.parents)]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            c.font = Font(size=9.5)
        r += 1
    wb_io = BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io


def class_schedule_workbook(school_class, grid, days, slots):
    wb = Workbook()
    ws = wb.active
    ws.title = "Emploi du temps"
    ncols = len(slots) + 1
    ws.column_dimensions["A"].width = 12
    for i in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    _header(ws, 1, ncols, f"LYCÉE TECHNIQUE DE TIBATI — EMPLOI DU TEMPS — {school_class.name}")
    r = 3
    ws.cell(row=r, column=1, value="Jour").font = Font(bold=True, color=NAVY)
    for i, (start, end) in enumerate(slots, start=2):
        c = ws.cell(row=r, column=i, value=f"{start} – {end}")
        c.font = Font(bold=True, size=9, color=NAVY)
        c.fill = PatternFill("solid", fgColor=CREAM)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    r += 1

    for day in days:
        ws.cell(row=r, column=1, value=day).font = Font(bold=True, size=10)
        ws.cell(row=r, column=1).border = BORDER
        entries_by_slot = {(e.start_time, e.end_time): e for e in grid.get(day, [])}
        for i, slot in enumerate(slots, start=2):
            e = entries_by_slot.get(slot)
            c = ws.cell(row=r, column=i)
            if e:
                c.value = f"{e.course.subject.name}\n{e.course.teacher.user.full_name}\n{e.room.name if e.room else 'Salle non définie'}"
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.font = Font(size=8.5)
        ws.row_dimensions[r].height = 40
        r += 1

    wb_io = BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io


def teacher_schedule_workbook(teacher, grid, days, slots, planned_slots=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Emploi du temps"
    ncols = len(slots) + 1
    ws.column_dimensions["A"].width = 12
    for i in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    _header(ws, 1, ncols, f"LYCÉE TECHNIQUE DE TIBATI — EMPLOI DU TEMPS INDIVIDUEL")
    planning_summary = f"Heures dues : {teacher.hours_due or 0} — Créneaux planifiés : {planned_slots if planned_slots is not None else 0} — Heures faites : {planned_slots if planned_slots is not None else 0}"
    ws.cell(row=2, column=1, value=f"Professeur : {teacher.user.full_name} — Département : {teacher.department.name if teacher.department else '—'} — {planning_summary}").font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    r = 4
    ws.cell(row=r, column=1, value="Jour").font = Font(bold=True, color=NAVY)
    for i, (start, end) in enumerate(slots, start=2):
        c = ws.cell(row=r, column=i, value=f"{start} – {end}")
        c.font = Font(bold=True, size=9, color=NAVY)
        c.fill = PatternFill("solid", fgColor=CREAM)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    r += 1

    for day in days:
        ws.cell(row=r, column=1, value=day).font = Font(bold=True, size=10)
        ws.cell(row=r, column=1).border = BORDER
        entries_by_slot = {(e.start_time, e.end_time): e for e in grid.get(day, [])}
        for i, slot in enumerate(slots, start=2):
            e = entries_by_slot.get(slot)
            c = ws.cell(row=r, column=i)
            if e:
                c.value = f"{e.course.subject.name}\n{e.course.school_class.code or e.course.school_class.name}\n{e.room.name if e.room else 'Salle non définie'}"
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.font = Font(size=8.5)
        ws.row_dimensions[r].height = 40
        r += 1

    wb_io = BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io


def honor_roll_register_workbook(rows, term, school_year):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tableaux d’honneur"
    for index, width in enumerate([7, 32, 18, 22, 18, 18, 14, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    _header(ws, 1, 9, "LYCÉE TECHNIQUE DE TIBATI — REGISTRE DES TABLEAUX D’HONNEUR")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    subtitle = ws.cell(row=2, column=1, value=f"{term} — Année scolaire {school_year} — Moyennes de 12/20 à 20/20")
    subtitle.font = Font(italic=True, size=10)
    subtitle.alignment = Alignment(horizontal="center")
    headers = ["N°", "Nom et prénom", "Date de naissance", "Lieu de naissance", "Matricule", "Classe", "Moyenne /20", "Rang", "Effectif"]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=column, value=header)
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=CREAM)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_number, item in enumerate(rows, start=5):
        school_class = item["school_class"]
        student = item["student"]
        values = [row_number - 4, student.full_name,
                  student.dob.strftime("%d/%m/%Y") if student.dob else "—", student.birth_place or "—",
                  student.matricule, school_class.name, item["average"], item["rank_label"], item["class_size"]]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_number, column=column, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if column == 7:
                cell.number_format = "0.00"
                cell.font = Font(bold=True, color=NAVY)
    if not rows:
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=9)
        ws.cell(row=5, column=1, value="Aucun élève admissible pour ce trimestre.").alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A5"
    wb_io = BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io
