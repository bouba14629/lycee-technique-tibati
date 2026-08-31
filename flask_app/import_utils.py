import csv
import re
import unicodedata
from datetime import datetime
from io import BytesIO, StringIO

from openpyxl import Workbook, load_workbook


MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 2000


def normalized_key(value):
    raw = str(value or "").strip().lower()
    without_accents = "".join(
        c for c in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(c)
    )
    return " ".join(re.sub(r"[^a-z0-9]+", " ", without_accents).split())


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def get_value(row, *names):
    for name in names:
        value = row.get(normalized_key(name))
        if value not in (None, ""):
            return cell_text(value)
    return ""


def parse_date(value):
    if not value:
        return None
    if hasattr(value, "date"):
        return value.date()
    text = cell_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("date invalide — utilisez AAAA-MM-JJ ou JJ/MM/AAAA")


def _is_blank_row(values):
    return not any(cell_text(value) for value in values)


def _csv_rows(content):
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1252")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\\t|")
        return list(csv.reader(StringIO(text), dialect))
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        return list(csv.reader(StringIO(text), delimiter=delimiter))


def _xlsx_rows(content):
    workbook = None
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if any(not _is_blank_row(row) for row in rows):
                return rows
    except Exception as exc:
        raise ValueError("fichier XLSX illisible ou corrompu") from exc
    finally:
        if workbook is not None:
            workbook.close()
    return []


def read_tabular_rows(file_storage):
    filename = (file_storage.filename or "").lower()
    if not filename.endswith((".csv", ".xlsx")):
        raise ValueError("format non pris en charge — utilisez un fichier CSV ou XLSX")
    content = file_storage.read()
    if not content:
        raise ValueError("fichier vide")
    if len(content) > MAX_IMPORT_BYTES:
        raise ValueError("fichier trop volumineux (5 Mo maximum)")

    rows = _csv_rows(content) if filename.endswith(".csv") else _xlsx_rows(content)
    header_index = next((index for index, row in enumerate(rows) if not _is_blank_row(row)), None)
    if header_index is None:
        raise ValueError("aucune ligne trouvée")
    raw_headers = rows[header_index]
    headers = [normalized_key(value) for value in raw_headers]
    non_empty_headers = [header for header in headers if header]
    if not non_empty_headers:
        raise ValueError("ligne d’en-tête introuvable")
    if len(non_empty_headers) != len(set(non_empty_headers)):
        raise ValueError("en-têtes en double — chaque colonne doit avoir un nom unique")

    parsed = []
    for line_number, values in enumerate(rows[header_index + 1:], start=header_index + 2):
        if _is_blank_row(values):
            continue
        row = {
            header: values[index] if index < len(values) else ""
            for index, header in enumerate(headers)
            if header
        }
        parsed.append((line_number, row))
        if len(parsed) > MAX_IMPORT_ROWS:
            raise ValueError(f"trop de lignes (maximum {MAX_IMPORT_ROWS})")
    if not parsed:
        raise ValueError("aucune ligne de données trouvée")
    return parsed


def import_template(kind):
    headers = {
        "enseignants": ["Nom complet", "Email", "Téléphone", "Département", "Spécialité", "Grade", "Heures dues"],
        "eleves": ["Nom", "Prénom", "Nom complet", "Matricule", "Classe", "Code classe", "Sexe", "Date de naissance", "Lieu de naissance", "Redoublant", "Statut"],
    }
    examples = {
        "enseignants": ["Marie NGONO", "marie.ngono@example.cm", "699000000", "ACA", "Bureautique", "PLET", 18],
        "eleves": ["MBOG", "Paul", "", "", "1A ACA", "ACA-1A", "M", "2008-09-14", "Tibati", "Non", "Inscrit"],
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import"
    sheet.append(headers[kind])
    sheet.append(examples[kind])
    for cell in sheet[1]:
        cell.font = cell.font.copy(bold=True, color="FFFFFF")
        cell.fill = cell.fill.copy(fgColor="0B2545", fill_type="solid")
    for col in sheet.columns:
        letter = col[0].column_letter
        sheet.column_dimensions[letter].width = max(16, min(28, max(len(cell_text(c.value)) for c in col) + 3))
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
