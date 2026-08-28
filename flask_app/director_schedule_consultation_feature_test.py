import os
import sys
from io import BytesIO

from openpyxl import load_workbook


os.environ["DATABASE_URL"] = "sqlite:////tmp/ltt-director-schedule-consultation.sqlite"
os.environ["LTT_ENV"] = "development"
os.environ["LTT_INITIAL_ADMIN_PASSWORD"] = "FoundateurTest#2026"

for module_name in [name for name in list(sys.modules)
                    if name in {"app", "models", "utils", "directeur_routes", "censeur_routes"}]:
    del sys.modules[module_name]

sys.path.insert(0, os.path.dirname(__file__))

from app import app
from models import Course, Department, ScheduleEntry, SchoolClass, Section, Subject, Teacher, User, db
from utils import build_official_grid, filled_official_slots


def main():
    app.config.update(TESTING=True)
    with app.app_context():
        db.drop_all()
        db.create_all()
        director = User(username="proviseur.test", role="directeur", full_name="Proviseur Test", must_change_password=False)
        director.set_password("Lyttib")
        censeur = User(username="censeur.test", role="censeur", full_name="Censeur Test", must_change_password=False)
        censeur.set_password("Lyttib")
        teacher_user = User(username="enseignant.test", role="enseignant", full_name="Enseignant Test", must_change_password=False)
        teacher_user.set_password("Lyttib")
        alpha_teacher_user = User(username="enseignant.alpha", role="enseignant", full_name="Alpha Enseignant", must_change_password=False)
        alpha_teacher_user.set_password("Lyttib")
        section = Section(name="Section Technique", code="STT")
        db.session.add_all([director, censeur, teacher_user, alpha_teacher_user, section])
        db.session.flush()
        department = Department(name="Électrotechnique", code="GEL", section_id=section.id)
        db.session.add(department)
        db.session.flush()
        school_class = SchoolClass(name="Zeta Classe", code="STT-2N-ELEC", level="2nde", department_id=department.id)
        alpha_class = SchoolClass(name="Alpha Classe", level="1A", department_id=department.id)
        subject = Subject(name="Électrotechnique", coefficient=2, department_id=department.id)
        teacher = Teacher(user_id=teacher_user.id, department_id=department.id, specialty="Électrotechnique", hours_due=3)
        alpha_teacher = Teacher(user_id=alpha_teacher_user.id, department_id=department.id, specialty="Automatisme")
        db.session.add_all([school_class, alpha_class, subject, teacher, alpha_teacher])
        db.session.flush()
        other_department = Department(name="Informatique", code="INFO", section_id=section.id)
        other_section = Section(name="Section Industrielle", code="IND")
        db.session.add_all([other_department, other_section])
        db.session.flush()
        department_class = SchoolClass(name="Delta Classe", level="P", department_id=other_department.id)
        other_department = Department(name="Électronique", code="ELEC", section_id=other_section.id)
        db.session.add(other_department)
        db.session.flush()
        other_section_class = SchoolClass(name="Omega Classe", level="Tle", department_id=other_department.id)
        db.session.add_all([department_class, other_section_class])
        course = Course(subject_id=subject.id, teacher_id=teacher.id, class_id=school_class.id)
        db.session.add(course)
        db.session.flush()
        db.session.add(ScheduleEntry(course_id=course.id, room_id=None, day="Lundi", start_time="08:20", end_time="09:10", published=True))
        db.session.add(ScheduleEntry(course_id=course.id, room_id=None, day="Mardi", start_time="10:15", end_time="11:55", published=True))
        db.session.commit()
        entries = ScheduleEntry.query.join(Course).filter(Course.teacher_id == teacher.id).all()
        assert filled_official_slots(build_official_grid(entries)) == 3
        ids = (school_class.id, teacher.id, section.id, department.id, department_class.id, other_section_class.id)

    class_id, teacher_id, section_id, department_id, department_class_id, other_section_class_id = ids
    with app.test_client() as client:
        assert client.post("/login", data={"username": "proviseur.test", "password": "Lyttib"}).status_code == 302
        class_overview = client.get("/directeur/emplois-du-temps?view=classe")
        assert class_overview.status_code == 200
        assert b"Emplois du temps par classe" in class_overview.data
        assert b"Emplois du temps individuels des enseignants" not in class_overview.data
        assert b'id="directorScheduleMenu"' in class_overview.data
        assert b"data-director-schedule-toggle" in class_overview.data
        assert b"?view=classe" in class_overview.data
        assert b"?view=individuel" in class_overview.data
        assert b"Enseignant Test" not in class_overview.data
        assert class_overview.data.index(b"Alpha Classe") < class_overview.data.index(b"Zeta Classe")
        assert b"Fili\xc3\xa8re / D\xc3\xa9partement" in class_overview.data
        assert b'id="schedule-section"' in class_overview.data
        assert b'id="schedule-department"' in class_overview.data
        assert b'id="schedule-class"' in class_overview.data

        section_filtered = client.get(f"/directeur/emplois-du-temps?view=classe&section_id={section_id}")
        assert section_filtered.status_code == 200
        assert f'data-schedule-class-id="{class_id}"'.encode() in section_filtered.data
        assert f'data-schedule-class-id="{department_class_id}"'.encode() in section_filtered.data
        assert f'data-schedule-class-id="{other_section_class_id}"'.encode() not in section_filtered.data

        department_filtered = client.get(f"/directeur/emplois-du-temps?view=classe&section_id={section_id}&department_id={department_id}")
        assert department_filtered.status_code == 200
        assert f'data-schedule-class-id="{class_id}"'.encode() in department_filtered.data
        assert f'data-schedule-class-id="{department_class_id}"'.encode() not in department_filtered.data

        class_filtered = client.get(f"/directeur/emplois-du-temps?view=classe&section_id={section_id}&department_id={department_id}&class_id={class_id}")
        assert class_filtered.status_code == 200
        assert f'data-schedule-class-id="{class_id}"'.encode() in class_filtered.data
        assert b'data-schedule-class-id="' not in class_filtered.data.replace(f'data-schedule-class-id="{class_id}"'.encode(), b"")

        teacher_overview = client.get("/directeur/emplois-du-temps?view=individuel")
        assert teacher_overview.status_code == 200
        assert b"Emplois du temps individuels des enseignants" in teacher_overview.data
        assert b"Emplois du temps par classe" not in teacher_overview.data
        assert b"Enseignant Test" in teacher_overview.data
        assert b"Alpha Enseignant" in teacher_overview.data
        assert teacher_overview.data.index(b"Alpha Enseignant") < teacher_overview.data.index(b"Enseignant Test")
        assert b'id="teacher-schedule-search"' in teacher_overview.data
        assert b"Cr\xc3\xa9neaux planifi\xc3\xa9s" not in teacher_overview.data
        assert b"Heures faites" in teacher_overview.data
        assert b"Heures dues" in teacher_overview.data
        teacher_row = teacher_overview.data.split(f'data-teacher-id="{teacher_id}"'.encode(), 1)[1].split(b"</tr>", 1)[0]
        assert b">3</td>" in teacher_row
        assert teacher_row.count(b">3</td>") == 2

        teacher_search = client.get("/directeur/emplois-du-temps?view=individuel&q=alpha")
        assert teacher_search.status_code == 200
        assert b'value="alpha"' in teacher_search.data
        assert b"Alpha Enseignant" in teacher_search.data
        assert b"Enseignant Test" not in teacher_search.data

        fallback_overview = client.get("/directeur/emplois-du-temps?view=invalide")
        assert fallback_overview.status_code == 200
        assert b"Emplois du temps par classe" in fallback_overview.data
        assert b"Emplois du temps individuels des enseignants" not in fallback_overview.data
        class_preview = client.get(f"/censeur/emplois-du-temps/{class_id}/officiel")
        assert class_preview.status_code == 200
        teacher_preview = client.get(f"/directeur/emplois-du-temps/enseignants/{teacher_id}")
        assert teacher_preview.status_code == 200
        assert b"EMPLOI DE TEMPS INDIVIDUEL" in teacher_preview.data
        assert b"STT-2N-ELEC" in teacher_preview.data
        assert b"Zeta Classe" not in teacher_preview.data
        assert b'HEURES DUES :</strong> <span style="color:#c0392b; font-weight:700;">3</span>' in teacher_preview.data
        assert b'CR\xc3\x89NEAUX PLANIFI\xc3\x89S :</strong> <span style="color:#c0392b; font-weight:700;">3</span>' in teacher_preview.data
        assert b'HEURES FAITES :</strong> <span style="color:#c0392b; font-weight:700;">3</span>' in teacher_preview.data
        teacher_pdf = client.get(f"/directeur/emplois-du-temps/enseignants/{teacher_id}/officiel.pdf")
        assert teacher_pdf.status_code == 200
        assert "application/pdf" in teacher_pdf.content_type
        teacher_export = client.get(f"/directeur/emplois-du-temps/enseignants/{teacher_id}/officiel.xlsx")
        assert teacher_export.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in teacher_export.content_type
        workbook = load_workbook(BytesIO(teacher_export.data))
        assert any("STT-2N-ELEC" in str(cell.value or "") for row in workbook.active.iter_rows() for cell in row)
        assert any("Heures dues : 3 — Créneaux planifiés : 3 — Heures faites : 3" in str(cell.value or "") for row in workbook.active.iter_rows() for cell in row)

        client.get("/logout")
        assert client.post("/login", data={"username": "censeur.test", "password": "Lyttib"}).status_code == 302
        censeur_overview = client.get("/censeur/emplois-du-temps/enseignants")
        assert censeur_overview.status_code == 200
        assert b"Emplois du temps individuels" in censeur_overview.data
        assert b"Imprimer (PDF)" in censeur_overview.data
        assert b"Enseignant Test" in censeur_overview.data
        censeur_preview = client.get(f"/censeur/emplois-du-temps/enseignants/{teacher_id}")
        assert censeur_preview.status_code == 200
        assert b"EMPLOI DE TEMPS INDIVIDUEL" in censeur_preview.data
        assert b"STT-2N-ELEC" in censeur_preview.data
        assert b"Zeta Classe" not in censeur_preview.data
        assert b'HEURES FAITES :</strong> <span style="color:#c0392b; font-weight:700;">3</span>' in censeur_preview.data
        censeur_pdf = client.get(f"/censeur/emplois-du-temps/enseignants/{teacher_id}/officiel.pdf")
        assert censeur_pdf.status_code == 200
        assert "application/pdf" in censeur_pdf.content_type

    assert app.config["TEMPLATES_AUTO_RELOAD"] is True
    assert app.jinja_env.auto_reload is True
    print("DIRECTOR_SCHEDULE_CONSULTATION_FEATURE_TEST_OK")


if __name__ == "__main__":
    main()
