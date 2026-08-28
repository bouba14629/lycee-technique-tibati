import random
import os
import json
import csv
import re
import unicodedata
from datetime import date, timedelta
from urllib.request import Request, urlopen
from flask import render_template, request, redirect, url_for, flash, session, abort
from app import app, db
from models import (
    User, Section, Department, SchoolClass, Subject, Teacher, Parent, Student,
    Room, Equipment, MaintenanceRequest, ScheduleEntry, Course, Reservation,
)
from utils import roles_required, notify, user_scoped_department_ids, generate_account_password, check_schedule_conflict


CLASS_LEVEL_LABELS = {
    "1A": "PREMIERE ANNEE",
    "2A": "DEUXIEME ANNEE",
    "3A": "TROISIEME ANNEE",
    "4A": "QUATRIEME ANNEE",
    "2nde": "SECONDE",
    "P": "PREMIERE",
    "Tle": "TERMINALE",
}


def normalize_import_header(value):
    """Uniformise les en-têtes de fichiers importés malgré les accents et variantes usuelles."""
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    without_accents = "".join(char for char in normalized if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", without_accents.replace("_", " ").strip().lower())


def teacher_name_from_import_row(data):
    """Accepte un nom complet ou l’association de colonnes Nom et Prénom."""
    complete_name = data.get("nom complet") or data.get("noms et prenoms") or data.get("nom et prenom")
    if complete_name:
        return " ".join(str(complete_name).split())
    last_name = data.get("nom") or data.get("nom de famille") or data.get("noms") or ""
    first_name = data.get("prenom") or data.get("prenoms") or data.get("prenom(s)") or ""
    return " ".join(part for part in (str(last_name).strip(), str(first_name).strip()) if part)


def slugify(s):
    return s.lower().replace(" ", ".").replace("'", "")


def gen_username(full_name):
    base = slugify(full_name)
    uname = base
    i = 1
    while User.query.filter_by(username=uname).first():
        i += 1
        uname = f"{base}{i}"
    return uname


UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "uploads", "students")
ALLOWED_PHOTO_EXT = {"jpg", "jpeg", "png", "webp"}


def save_student_photo(file_storage, matricule):
    """Enregistre une photo dans le stockage persistant WebDev ; renvoie son URL ou None."""
    if not file_storage or not file_storage.filename:
        return None
    ext = file_storage.filename.rsplit(".", 1)[-1].lower() if "." in file_storage.filename else ""
    if ext not in ALLOWED_PHOTO_EXT:
        return None
    file_bytes = file_storage.read()
    if not file_bytes:
        return None
    content_type = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}[ext]
    upload_port = os.getenv("PORT", "3000")
    upload_request = Request(
        f"http://127.0.0.1:{upload_port}/api/ltt/media",
        data=file_bytes,
        method="POST",
        headers={
            "Content-Type": "application/octet-stream",
            "X-LTT-Internal-Token": os.getenv("JWT_SECRET", ""),
            "X-File-Name": f"{matricule}.{ext}",
            "X-File-Type": content_type,
        },
    )
    try:
        with urlopen(upload_request, timeout=20) as response:
            return json.loads(response.read().decode("utf-8"))["url"]
    except Exception:
        return None


# ----------------------------------------------------------- utilisateurs ---
@app.route("/directeur/utilisateurs")
@roles_required("directeur", "censeur", "conseiller_orientation")
def dir_users():
    from models import STAFF_GRADES
    role_filter = request.args.get("role", "")
    search = request.args.get("q", "").strip()
    q = User.query
    if role_filter:
        q = q.filter_by(role=role_filter)
    if search:
        like = f"%{search}%"
        q = q.filter(db.or_(User.full_name.ilike(like), User.username.ilike(like), User.email.ilike(like)))
    sort = request.args.get("sort", "full_name")
    sort_dir = request.args.get("dir", "asc")
    sort_col = {"full_name": User.full_name, "username": User.username, "role": User.role}.get(sort, User.full_name)
    sort_col = sort_col.desc() if sort_dir == "desc" else sort_col.asc()
    users = q.order_by(sort_col).all()
    sections = Section.query.order_by(Section.name).all()
    return render_template("dir_users.html", users=users, role_filter=role_filter, search=search,
                            sort=sort, sort_dir=sort_dir,
                            sections=sections, staff_grades=STAFF_GRADES)


@app.route("/directeur/utilisateurs/export.xlsx")
@roles_required("directeur", "censeur")
def dir_users_export():
    from flask import send_file
    from excel_utils import users_workbook
    role_filter = request.args.get("role", "")
    search = request.args.get("q", "").strip()
    q = User.query
    if role_filter:
        q = q.filter_by(role=role_filter)
    if search:
        like = f"%{search}%"
        q = q.filter(db.or_(User.full_name.ilike(like), User.username.ilike(like), User.email.ilike(like)))
    users = q.order_by(User.role, User.full_name).all()
    from models import ROLE_LABELS
    title = f"LTT — {ROLE_LABELS.get(role_filter, 'Utilisateurs')}" if role_filter else "LTT — Tous les utilisateurs"
    wb_io = users_workbook(users, title, role_filter)
    filename = f"LTT_{role_filter or 'utilisateurs'}.xlsx"
    return send_file(wb_io, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      as_attachment=True, download_name=filename)


@app.route("/directeur/utilisateurs/nouveau", methods=["GET", "POST"])
@roles_required("directeur")
def dir_user_new():
    departments = Department.query.order_by(Department.name).all()
    sections = Section.query.order_by(Section.name).all()
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        role = request.form.get("role")
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        civility = request.form.get("civility", "").strip()
        section_id = request.form.get("section_id", type=int)
        grade = request.form.get("grade", "").strip()
        valid_roles = ("directeur", "censeur", "censeur_crm", "surveillant_general", "conseiller_orientation",
                        "chef_travaux", "chef_crm", "enseignant")
        if not full_name or role not in valid_roles:
            flash("Nom complet et rôle valides requis (les élèves/parents sont créés via Inscription).", "warning")
            return redirect(url_for("dir_user_new"))
        if civility and civility not in {"Mme.", "M."}:
            flash("La civilité doit être Mme. ou M.", "warning")
            return redirect(url_for("dir_user_new"))
        uname = gen_username(full_name)
        temp_pw = generate_account_password(full_name, role)
        u = User(username=uname, role=role, full_name=full_name, email=email, phone=phone, civility=civility or None,
                 must_change_password=True,
                 section_id=section_id if role in ("censeur", "surveillant_general", "chef_travaux") else None,
                 grade=grade if role != "enseignant" else None)
        u.set_password(temp_pw)
        db.session.add(u)
        db.session.flush()
        if role == "enseignant":
            dept_id = request.form.get("department_id", type=int)
            t = Teacher(user_id=u.id, specialty=request.form.get("specialty", ""), department_id=dept_id,
                        grade=grade,
                        hours_due=request.form.get("hours_due", 18, type=int),
                        hire_date=date.today())
            db.session.add(t)
        db.session.commit()
        flash(f"Compte créé — identifiant : {uname} — mot de passe : {temp_pw}", "success")
        return redirect(url_for("dir_users"))
    return render_template("dir_user_new.html", departments=departments, sections=sections)


@app.route("/directeur/utilisateurs/import", methods=["POST"])
@roles_required("directeur")
def teachers_import():
    from openpyxl import load_workbook
    file = request.files.get("import_file")
    if not file or not file.filename:
        flash("Veuillez sélectionner un fichier Excel (.xlsx).", "warning")
        return redirect(url_for("dir_users"))
    filename = file.filename.lower()
    rows = []
    try:
        if filename.endswith(".csv"):
            raw_csv = file.read().decode("utf-8-sig")
            rows = list(csv.reader(raw_csv.splitlines()))
        elif filename.endswith(".xlsx"):
            wb = load_workbook(file, data_only=True)
            rows = list(wb.active.iter_rows(min_row=1, values_only=True))
        else:
            flash("Format non pris en charge — utilisez un fichier CSV ou XLSX.", "danger")
            return redirect(url_for("dir_users"))
    except UnicodeDecodeError:
        flash("Le fichier CSV doit être encodé en UTF-8.", "danger")
        return redirect(url_for("dir_users"))
    except Exception:
        flash("Fichier illisible — vérifiez le format CSV ou XLSX.", "danger")
        return redirect(url_for("dir_users"))
    if not rows:
        flash("Le fichier sélectionné est vide.", "warning")
        return redirect(url_for("dir_users"))

    depts_by_code = {d.code.strip().lower(): d for d in Department.query.all()}
    default_department = Department.query.get(request.form.get("department_id", type=int))
    created, skipped = 0, 0
    skipped_reasons = []

    header_row = [normalize_import_header(value) for value in rows[0]]
    for row in rows[1:]:
        if not any(row):
            continue
        data = dict(zip(header_row, row))
        full_name = teacher_name_from_import_row(data)
        if not full_name:
            skipped += 1
            skipped_reasons.append("Nom complet absent (utilisez « Nom complet » ou « Nom » et « Prénom »).")
            continue
        email = str(data.get("email") or "").strip()
        phone = str(data.get("telephone") or data.get("tel") or data.get("phone") or "").strip()
        dept_code = str(data.get("departement") or data.get("filiere") or "").strip()
        specialty = str(data.get("specialite") or data.get("matiere") or "").strip()
        grade = str(data.get("grade") or "").strip()
        hours_due = data.get("heures dues") or data.get("heures") or 18
        try:
            hours_due = int(hours_due)
        except (TypeError, ValueError):
            hours_due = 18
        dept = depts_by_code.get(dept_code.lower()) or default_department
        if dept_code and not dept:
            skipped += 1
            skipped_reasons.append(f"Département introuvable : « {dept_code} » ({full_name})")
            continue

        uname = gen_username(full_name)
        temp_pw = f"LTT{random.randint(1000,9999)}!"
        u = User(username=uname, role="enseignant", full_name=full_name, email=email, phone=phone,
                 must_change_password=True)
        u.set_password(temp_pw)
        db.session.add(u)
        db.session.flush()
        t = Teacher(user_id=u.id, specialty=specialty or (dept.name if dept else ""),
                    department_id=dept.id if dept else None, grade=grade,
                    hours_due=hours_due, hire_date=date.today())
        db.session.add(t)
        created += 1

    db.session.commit()
    msg = f"{created} enseignant(s) importé(s) avec succès."
    if skipped:
        msg += f" {skipped} ligne(s) ignorée(s)."
    flash(msg, "success" if created else "warning")
    if skipped_reasons:
        flash(" · ".join(skipped_reasons[:5]) + (" …" if len(skipped_reasons) > 5 else ""), "warning")
    return redirect(url_for("dir_users"))


@app.route("/directeur/utilisateurs/<int:user_id>/basculer")
@roles_required("directeur")
def dir_user_toggle(user_id):
    u = User.query.get_or_404(user_id)
    u.active = not u.active
    db.session.commit()
    flash(f"Compte {u.username} {'activé' if u.active else 'désactivé'}.", "info")
    return redirect(url_for("dir_users"))


@app.route("/directeur/utilisateurs/<int:user_id>/reinitialiser", methods=["POST"])
@roles_required("directeur")
def dir_user_reset(user_id):
    u = User.query.get_or_404(user_id)
    temp_pw = generate_account_password(u.full_name, u.role)
    u.set_password(temp_pw)
    u.must_change_password = True
    db.session.commit()
    flash(f"Nouveau mot de passe temporaire généré pour {u.username} : {temp_pw}. L’utilisateur devra le remplacer à sa prochaine connexion.", "success")
    return redirect(url_for("dir_users"))


@app.route("/directeur/utilisateurs/<int:user_id>/modifier", methods=["POST"])
@roles_required("directeur")
def dir_user_edit(user_id):
    u = User.query.get_or_404(user_id)
    u.full_name = request.form.get("full_name", u.full_name).strip()
    u.email = request.form.get("email", u.email or "").strip()
    u.phone = request.form.get("phone", u.phone or "").strip()
    section_id = request.form.get("section_id", type=int)
    if u.role in ("censeur", "surveillant_general", "chef_travaux"):
        u.section_id = section_id
    grade = request.form.get("grade", "")
    if u.role == "enseignant" and u.teacher_profile:
        u.teacher_profile.specialty = request.form.get("specialty", u.teacher_profile.specialty)
        u.teacher_profile.grade = grade
        u.teacher_profile.hours_due = request.form.get("hours_due", u.teacher_profile.hours_due, type=int)
    elif grade:
        u.grade = grade
    new_pw = request.form.get("new_password", "").strip()
    if new_pw:
        u.set_password(new_pw)
        u.must_change_password = False
    db.session.commit()
    flash(f"Compte {u.username} modifié.", "success")
    return redirect(url_for("dir_users", role=request.form.get("role_filter", "")))


@app.route("/directeur/utilisateurs/<int:user_id>/supprimer")
@roles_required("directeur")
def dir_user_delete(user_id):
    u = User.query.get_or_404(user_id)
    if u.role in ("eleve", "parent"):
        flash("Utilisez la fiche de l'élève pour supprimer un compte élève ou parent.", "warning")
        return redirect(url_for("dir_users", role=u.role))
    if u.role == "directeur" and User.query.filter_by(role="directeur").count() <= 1:
        flash("Impossible de supprimer le dernier compte Proviseur.", "danger")
        return redirect(url_for("dir_users"))
    if u.id == session.get("user_id"):
        flash("Vous ne pouvez pas supprimer votre propre compte.", "danger")
        return redirect(url_for("dir_users"))
    if u.teacher_profile and u.teacher_profile.courses:
        flash(f"Impossible de supprimer « {u.full_name} » : {len(u.teacher_profile.courses)} cours lui sont affectés. Réaffectez-les d'abord.", "danger")
        return redirect(url_for("dir_users"))
    uname = u.username
    if u.teacher_profile:
        db.session.delete(u.teacher_profile)
    db.session.delete(u)
    db.session.commit()
    flash(f"Compte « {uname} » supprimé.", "info")
    return redirect(url_for("dir_users"))


# -------------------------------------------------------- inscription élèves ---
@app.route("/parents/nouveau", methods=["GET", "POST"])
@roles_required("directeur", "conseiller_orientation")
def parent_new():
    search = request.args.get("q", "").strip()
    students = []
    if search:
        like = f"%{search}%"
        students = (Student.query.filter(db.or_(Student.first_name.ilike(like), Student.last_name.ilike(like),
                                                  Student.matricule.ilike(like)))
                    .order_by(Student.last_name).limit(20).all())
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        phone = request.form.get("phone", "").strip()
        profession = request.form.get("profession", "").strip()
        student_ids = request.form.getlist("student_ids", type=int)
        if not full_name or not student_ids:
            flash("Nom du parent et au moins un enfant sont obligatoires.", "warning")
            return redirect(url_for("parent_new"))
        uname = gen_username(full_name)
        temp_pw = generate_account_password(full_name, "parent")
        pu = User(username=uname, role="parent", full_name=full_name, phone=phone,
                  must_change_password=True)
        pu.set_password(temp_pw)
        db.session.add(pu)
        db.session.flush()
        pp = Parent(user_id=pu.id, phone=phone, profession=profession)
        db.session.add(pp)
        db.session.flush()
        for sid in student_ids:
            st = Student.query.get(sid)
            if st:
                pp.children.append(st)
        db.session.commit()
        flash(f"Compte parent créé — identifiant : {uname} — mot de passe : {temp_pw}", "success")
        return redirect(url_for("dir_users", role="parent"))
    return render_template("parent_new.html", search=search, students=students)


@app.route("/eleves/inscription", methods=["GET", "POST"])
@roles_required("directeur")
def student_enroll():
    classes = SchoolClass.query.join(Department).order_by(Department.name, SchoolClass.level).all()
    existing_parents = Parent.query.join(User).order_by(User.full_name).all()
    if request.method == "POST":
        first = request.form.get("first_name", "").strip()
        last = request.form.get("last_name", "").strip()
        matricule = request.form.get("matricule", "").strip().upper()
        sex = request.form.get("sex")
        dob = request.form.get("dob")
        address = request.form.get("birth_place", "").strip()
        class_id = request.form.get("class_id", type=int)
        existing_parent_id = request.form.get("existing_parent_id", type=int)
        parent_name = request.form.get("parent_name", "").strip()
        parent_phone = request.form.get("parent_phone", "").strip()
        parent_profession = request.form.get("parent_profession", "").strip()
        reinscription = request.form.get("status") == "Réinscrit"
        is_repeater = request.form.get("is_repeater") == "1"

        if not (first and last and matricule and class_id):
            flash("Nom, prénom, matricule et classe sont obligatoires.", "warning")
            return redirect(url_for("student_enroll"))
        if len(matricule) > 30:
            flash("Le matricule ne peut pas dépasser 30 caractères.", "warning")
            return redirect(url_for("student_enroll"))
        if Student.query.filter_by(matricule=matricule).first():
            flash(f"Le matricule « {matricule} » est déjà utilisé par un autre élève.", "danger")
            return redirect(url_for("student_enroll"))

        # Compte de connexion de l'élève — exclusif, pour qu'il consulte uniquement ses propres informations
        student_uname = gen_username(f"{first} {last}")
        student_temp_pw = generate_account_password(f"{first} {last}", "eleve")
        su = User(username=student_uname, role="eleve", full_name=f"{first} {last}",
                  must_change_password=True)
        su.set_password(student_temp_pw)
        db.session.add(su)
        db.session.flush()

        student = Student(user_id=su.id, matricule=matricule, first_name=first, last_name=last, sex=sex,
                           dob=date.fromisoformat(dob) if dob else None, birth_place=address,
                           class_id=class_id, status="Réinscrit" if reinscription else "Inscrit",
                           is_repeater=is_repeater)
        db.session.add(student)
        db.session.flush()
        photo_file = request.files.get("photo")
        photo_name = save_student_photo(photo_file, matricule)
        if photo_name:
            student.photo = photo_name
            db.session.commit()

        msg = (f"Élève {student.full_name} inscrit (matricule {matricule}). "
               f"Compte élève : {student_uname} / {student_temp_pw}")

        if existing_parent_id:
            existing_parent = Parent.query.get(existing_parent_id)
            if existing_parent:
                existing_parent.children.append(student)
                msg += f" — Rattaché au compte parent existant : {existing_parent.user.username}"
        elif parent_name:
            uname = gen_username(parent_name)
            temp_pw = generate_account_password(parent_name, "parent")
            pu = User(username=uname, role="parent", full_name=parent_name,
                      must_change_password=True)
            pu.set_password(temp_pw)
            db.session.add(pu)
            db.session.flush()
            pp = Parent(user_id=pu.id, phone=parent_phone, profession=parent_profession)
            db.session.add(pp)
            db.session.flush()
            pp.children.append(student)
            msg += f" — Compte parent : {uname} / {temp_pw}"

        db.session.commit()
        flash(msg, "success")
        return redirect(url_for("students_list"))
    return render_template("student_enroll.html", classes=classes, existing_parents=existing_parents)


@app.route("/eleves/import", methods=["POST"])
@roles_required("directeur")
def students_import():
    from openpyxl import load_workbook
    file = request.files.get("import_file")
    if not file or not file.filename:
        flash("Veuillez sélectionner un fichier Excel (.xlsx).", "warning")
        return redirect(url_for("student_enroll"))
    try:
        wb = load_workbook(file, data_only=True)
    except Exception:
        flash("Fichier illisible — assurez-vous qu'il s'agit bien d'un fichier .xlsx.", "danger")
        return redirect(url_for("student_enroll"))
    ws = wb.active

    classes_by_name = {c.name.strip().lower(): c for c in SchoolClass.query.all()}
    created, skipped = 0, 0
    skipped_reasons = []
    year_prefix = date.today().year

    header_row = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if header_row is None:
            header_row = [str(v).strip().lower() if v else "" for v in row]
            continue
        if not any(row):
            continue
        data = dict(zip(header_row, row))
        full_name = str(data.get("nom complet") or "").strip()
        class_name = str(data.get("classe") or "").strip()
        if not full_name or not class_name:
            skipped += 1
            continue
        cls = classes_by_name.get(class_name.lower())
        if not cls:
            skipped += 1
            skipped_reasons.append(f"Classe introuvable : « {class_name} » ({full_name})")
            continue
        parts = full_name.split(" ", 1)
        first, last = (parts[0], parts[1]) if len(parts) > 1 else (parts[0], "")
        sexe = str(data.get("sexe") or "").strip().upper()[:1] or None
        adresse = str(data.get("lieu de naissance") or data.get("adresse") or "").strip()
        matricule = f"LTT{year_prefix}{random.randint(1000,9999)}"
        while Student.query.filter_by(matricule=matricule).first():
            matricule = f"LTT{year_prefix}{random.randint(1000,9999)}"

        student_uname = gen_username(full_name)
        student_temp_pw = f"LTT{random.randint(1000,9999)}!"
        su = User(username=student_uname, role="eleve", full_name=full_name, must_change_password=True)
        su.set_password(student_temp_pw)
        db.session.add(su)
        db.session.flush()
        student = Student(user_id=su.id, matricule=matricule, first_name=first, last_name=last,
                           sex=sexe, address=adresse, class_id=cls.id, status="Inscrit")
        db.session.add(student)
        created += 1

    db.session.commit()
    msg = f"{created} élève(s) importé(s) avec succès."
    if skipped:
        msg += f" {skipped} ligne(s) ignorée(s) (classe introuvable ou données manquantes)."
    flash(msg, "success" if created else "warning")
    if skipped_reasons:
        flash(" · ".join(skipped_reasons[:5]) + (" …" if len(skipped_reasons) > 5 else ""), "warning")
    return redirect(url_for("students_list"))


def _teacher_class_ids():
    """Classes où l'enseignant connecté intervient (pour restreindre sa vue des élèves)."""
    user = User.query.get(session["user_id"])
    teacher = user.teacher_profile if user else None
    if not teacher:
        return set()
    return {c.class_id for c in teacher.courses}


def _assert_student_card_access(student):
    """Applique le périmètre élève déjà utilisé par le dossier aux cartes scolaires."""
    from utils import user_scoped_class_ids
    user = User.query.get(session["user_id"])
    if session.get("role") == "enseignant" and student.class_id not in _teacher_class_ids():
        abort(403)
    if session.get("role") == "surveillant_general":
        allowed_ids = user_scoped_class_ids(user)
        if allowed_ids is not None and student.class_id not in allowed_ids:
            abort(403)


@app.route("/eleves")
@roles_required("directeur", "censeur", "surveillant_general", "conseiller_orientation", "enseignant")
def students_list():
    from utils import user_scoped_class_ids
    class_id = request.args.get("class_id", type=int)
    search = request.args.get("q", "").strip()
    q = Student.query
    classes = SchoolClass.query.join(Department).order_by(Department.name, SchoolClass.level).all()
    user = User.query.get(session["user_id"])
    if session.get("role") == "enseignant":
        allowed_ids = _teacher_class_ids()
        classes = [c for c in classes if c.id in allowed_ids]
        q = q.filter(Student.class_id.in_(allowed_ids)) if allowed_ids else q.filter(db.false())
    elif session.get("role") == "surveillant_general":
        allowed_ids = user_scoped_class_ids(user)
        if allowed_ids is not None:
            classes = [c for c in classes if c.id in allowed_ids]
            q = q.filter(Student.class_id.in_(allowed_ids)) if allowed_ids else q.filter(db.false())
    if class_id:
        q = q.filter_by(class_id=class_id)
    if search:
        like = f"%{search}%"
        q = q.filter(db.or_(Student.first_name.ilike(like), Student.last_name.ilike(like), Student.matricule.ilike(like)))
    sort = request.args.get("sort", "last_name")
    sort_dir = request.args.get("dir", "asc")
    if sort == "last_name" and sort_dir == "asc":
        students = q.order_by(db.func.lower(Student.last_name), db.func.lower(Student.first_name)).all()
    else:
        sort_col = {"last_name": Student.last_name, "matricule": Student.matricule,
                    "dob": Student.dob, "birth_place": Student.birth_place}.get(sort, Student.last_name)
        sort_col = sort_col.desc() if sort_dir == "desc" else sort_col.asc()
        students = q.order_by(sort_col).all()
    return render_template("students_list.html", students=students, classes=classes, class_id=class_id,
                            search=search, sort=sort, sort_dir=sort_dir)


@app.route("/eleves/export.xlsx")
@roles_required("directeur", "censeur", "surveillant_general", "conseiller_orientation", "enseignant")
def students_export():
    from flask import send_file
    from excel_utils import students_workbook
    class_id = request.args.get("class_id", type=int)
    search = request.args.get("q", "").strip()
    q = Student.query
    if session.get("role") == "enseignant":
        allowed_ids = _teacher_class_ids()
        q = q.filter(Student.class_id.in_(allowed_ids)) if allowed_ids else q.filter(db.false())
    if class_id:
        q = q.filter_by(class_id=class_id)
    if search:
        like = f"%{search}%"
        q = q.filter(db.or_(Student.first_name.ilike(like), Student.last_name.ilike(like), Student.matricule.ilike(like)))
    students = q.order_by(db.func.lower(Student.last_name), db.func.lower(Student.first_name)).all()
    wb_io = students_workbook(students, "LYCÉE TECHNIQUE DE TIBATI — LISTE DES ÉLÈVES")
    return send_file(wb_io, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      as_attachment=True, download_name="LTT_eleves.xlsx")


@app.route("/eleves/export.pdf")
@roles_required("directeur", "censeur", "surveillant_general", "conseiller_orientation", "enseignant")
def students_export_pdf():
    from flask import send_file, abort
    from pdf_utils import render_pdf
    class_id = request.args.get("class_id", type=int)
    search = request.args.get("q", "").strip()
    q = Student.query
    if session.get("role") == "enseignant":
        allowed_ids = _teacher_class_ids()
        q = q.filter(Student.class_id.in_(allowed_ids)) if allowed_ids else q.filter(db.false())
    if class_id:
        q = q.filter_by(class_id=class_id)
    if search:
        like = f"%{search}%"
        q = q.filter(db.or_(Student.first_name.ilike(like), Student.last_name.ilike(like), Student.matricule.ilike(like)))
    students = q.order_by(db.func.lower(Student.last_name), db.func.lower(Student.first_name)).all()
    pdf = render_pdf("pdf/students_list_pdf.html", students=students)
    if not pdf:
        abort(500)
    return send_file(pdf, mimetype="application/pdf", as_attachment=True, download_name="LTT_eleves.pdf")


@app.route("/eleves/<int:student_id>")
@roles_required("directeur", "censeur", "surveillant_general", "conseiller_orientation", "enseignant")
def student_detail(student_id):
    from utils import general_average, subject_averages, user_scoped_class_ids
    student = Student.query.get_or_404(student_id)
    user = User.query.get(session["user_id"])
    if session.get("role") == "enseignant" and student.class_id not in _teacher_class_ids():
        abort(403)
    if session.get("role") == "surveillant_general":
        allowed_ids = user_scoped_class_ids(user)
        if allowed_ids is not None and student.class_id not in allowed_ids:
            abort(403)
    avg = general_average(student.id)
    subj_avgs = subject_averages(student.id)
    return render_template("student_detail.html", student=student, avg=avg, subj_avgs=subj_avgs)


@app.route("/eleves/<int:student_id>/carte")
@roles_required("directeur", "censeur", "surveillant_general", "conseiller_orientation", "enseignant")
def student_card_preview(student_id):
    from app import student_photo_url
    from student_card_utils import card_qr_data_uri, card_school_year, card_validity
    student = Student.query.get_or_404(student_id)
    _assert_student_card_access(student)
    return render_template("student_card.html", student=student, photo_url=student_photo_url(student.photo),
                           qr_data_uri=card_qr_data_uri(student), school_year=card_school_year(student),
                           validity=card_validity(student))


@app.route("/eleves/<int:student_id>/carte/telecharger.pdf")
@roles_required("directeur", "censeur", "surveillant_general", "conseiller_orientation", "enseignant")
def student_card_pdf(student_id):
    from flask import send_file
    from app import LTT_ASSETS
    from pdf_utils import pdf_asset, render_pdf
    from student_card_utils import card_qr_file, card_school_year, card_validity
    student = Student.query.get_or_404(student_id)
    _assert_student_card_access(student)
    storage_photo = student.photo if student.photo and student.photo.startswith("/manus-storage/") else LTT_ASSETS["img/avatar_placeholder.png"]
    photo_path = pdf_asset(("img", "avatar_placeholder.png"), storage_photo)
    logo_path = pdf_asset(("img", "logo.png"), LTT_ASSETS["img/logo.png"])
    photo_path = photo_path if os.path.exists(photo_path) else None
    logo_path = logo_path if os.path.exists(logo_path) else None
    pdf = render_pdf("pdf/student_card_pdf.html", student=student, photo_path=photo_path, logo_path=logo_path,
                     qr_path=card_qr_file(student), school_year=card_school_year(student), validity=card_validity(student))
    if not pdf:
        abort(500)
    filename = f"Carte_scolaire_{student.matricule}.pdf".replace(" ", "_")
    return send_file(pdf, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/cartes-scolaires/verifier/<token>")
def student_card_verify(token):
    from student_card_utils import validate_card_token, card_school_year, card_validity
    payload = validate_card_token(token)
    student = Student.query.get(payload["student_id"]) if payload else None
    valid = bool(student and student.status in {"Inscrit", "Réinscrit"} and payload["school_year"] == card_school_year(student))
    return render_template("student_card_verify.html", valid=valid, student=student if valid else None,
                           school_year=card_school_year(student) if valid else None,
                           validity=card_validity(student) if valid else None)


@app.route("/eleves/<int:student_id>/modifier", methods=["POST"])
@roles_required("directeur")
def student_edit(student_id):
    student = Student.query.get_or_404(student_id)
    student.first_name = request.form.get("first_name", student.first_name).strip()
    student.last_name = request.form.get("last_name", student.last_name).strip()
    student.sex = request.form.get("sex", student.sex)
    student.birth_place = request.form.get("birth_place", student.birth_place)
    student.is_repeater = request.form.get("is_repeater") == "1"
    new_matricule = request.form.get("matricule", "").strip()
    if new_matricule and new_matricule != student.matricule:
        if Student.query.filter(Student.matricule == new_matricule, Student.id != student.id).first():
            flash(f"Le matricule « {new_matricule} » est déjà utilisé par un autre élève.", "danger")
            return redirect(url_for("student_detail", student_id=student.id))
        student.matricule = new_matricule
    class_id = request.form.get("class_id", type=int)
    if class_id:
        student.class_id = class_id
    photo_file = request.files.get("photo")
    photo_name = save_student_photo(photo_file, student.matricule)
    if photo_name:
        student.photo = photo_name
    new_pw = request.form.get("new_password", "").strip()
    if new_pw and student.user:
        student.user.set_password(new_pw)
        student.user.must_change_password = False
    if student.user:
        student.user.full_name = f"{student.first_name} {student.last_name}"
    db.session.commit()
    flash("Fiche élève modifiée.", "success")
    return redirect(url_for("student_detail", student_id=student.id))


@app.route("/eleves/<int:student_id>/supprimer")
@roles_required("directeur")
def student_delete(student_id):
    from models import Grade, Attendance, Sanction, Reward
    student = Student.query.get_or_404(student_id)
    name = student.full_name
    Grade.query.filter_by(student_id=student.id).delete()
    Attendance.query.filter_by(student_id=student.id).delete()
    Sanction.query.filter_by(student_id=student.id).delete()
    Reward.query.filter_by(student_id=student.id).delete()
    user = student.user
    db.session.delete(student)
    if user:
        db.session.delete(user)
    db.session.commit()
    flash(f"Élève « {name} » et toutes ses données ont été supprimés.", "info")
    return redirect(url_for("students_list"))


@app.route("/eleves/<int:student_id>/sanction", methods=["POST"])
@roles_required("surveillant_general")
def student_sanction(student_id):
    from models import Sanction
    from utils import user_scoped_class_ids
    student = Student.query.get_or_404(student_id)
    user = User.query.get(session["user_id"])
    allowed_ids = user_scoped_class_ids(user)
    if allowed_ids is not None and student.class_id not in allowed_ids:
        abort(403)
    db.session.add(Sanction(student_id=student.id, type=request.form.get("type"),
                             description=request.form.get("description", ""),
                             date=date.today(), issued_by_id=session["user_id"]))
    if student.user_id:
        notify(student.user_id, f"Une sanction a été enregistrée dans votre dossier ({request.form.get('type')}).")
    for p in student.parents:
        notify(p.user_id, f"Une sanction a été enregistrée pour {student.full_name} ({request.form.get('type')}).")
    db.session.commit()
    flash("Sanction enregistrée.", "info")
    return redirect(url_for("student_detail", student_id=student.id))


@app.route("/eleves/<int:student_id>/recompense", methods=["POST"])
@roles_required("surveillant_general")
def student_reward(student_id):
    from models import Reward
    from utils import user_scoped_class_ids
    student = Student.query.get_or_404(student_id)
    user = User.query.get(session["user_id"])
    allowed_ids = user_scoped_class_ids(user)
    if allowed_ids is not None and student.class_id not in allowed_ids:
        abort(403)
    db.session.add(Reward(student_id=student.id, description=request.form.get("description", ""),
                           date=date.today(), issued_by_id=session["user_id"]))
    for p in student.parents:
        notify(p.user_id, f"{student.full_name} a reçu une distinction : {request.form.get('description', '')}")
    db.session.commit()
    flash("Récompense enregistrée.", "success")
    return redirect(url_for("student_detail", student_id=student.id))


# --------------------------------------------------- structure pédagogique ---
@app.route("/directeur/structure/classe/<int:class_id>/professeur-principal", methods=["POST"])
@roles_required("directeur", "censeur")
def dir_class_homeroom(class_id):
    cls = SchoolClass.query.get_or_404(class_id)
    user = User.query.get(session["user_id"])
    scoped_dept_ids = user_scoped_department_ids(user) if user.role == "censeur" else None
    if scoped_dept_ids is not None and cls.department_id not in scoped_dept_ids:
        abort(403)
    teacher_id = request.form.get("teacher_id", type=int)
    teacher = Teacher.query.get(teacher_id) if teacher_id else None
    if teacher and teacher.department_id != cls.department_id:
        flash("Le professeur principal doit appartenir au département de cette classe.", "danger")
        return redirect(url_for("dir_structure"))
    cls.homeroom_teacher_id = teacher.id if teacher else None
    db.session.commit()
    flash(f"Professeur principal de {cls.name} mis à jour.", "success")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure")
@roles_required("directeur", "censeur", "censeur_crm")
def dir_structure():
    from utils import user_scoped_department_ids
    user = User.query.get(session["user_id"])
    scoped_dept_ids = user_scoped_department_ids(user)
    if scoped_dept_ids is not None and user.section_id:
        # Censeur STT ou Industriel : uniquement sa propre section
        sections = Section.query.filter_by(id=user.section_id).all()
    else:
        # Proviseur ou Censeur Enseignements Généraux (portée transversale) : les deux sections
        sections = Section.query.order_by(Section.name).all()
    last_subject_class_id = None
    if user.role == "censeur":
        saved_class_id = session.get("last_subject_class_id")
        available_class_ids = {
            school_class.id
            for section in sections
            for department in section.departments
            for school_class in department.classes
        }
        if saved_class_id in available_class_ids:
            last_subject_class_id = saved_class_id
        elif saved_class_id is not None:
            session.pop("last_subject_class_id", None)
    return render_template("dir_structure.html", sections=sections, scoped_dept_ids=scoped_dept_ids,
                           last_subject_class_id=last_subject_class_id)


@app.route("/directeur/structure/section/nouvelle", methods=["POST"])
@roles_required("directeur")
def dir_section_new():
    name = request.form.get("name", "").strip()
    code = request.form.get("code", "").strip().upper()
    if not name or not code:
        flash("Le nom et le code de la section sont requis.", "warning")
    elif len(code) > 20:
        flash("Le code de section ne peut pas dépasser 20 caractères.", "warning")
    elif Section.query.filter((Section.name == name) | (Section.code == code)).first():
        flash("Une section avec ce nom ou ce code existe déjà.", "warning")
    else:
        db.session.add(Section(name=name, code=code))
        db.session.commit()
        flash("Section ajoutée. Vous pouvez maintenant créer ses filières.", "success")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure/section/<int:section_id>/modifier", methods=["POST"])
@roles_required("directeur")
def dir_section_edit(section_id):
    section = Section.query.get_or_404(section_id)
    name = request.form.get("name", "").strip()
    code = request.form.get("code", "").strip().upper()
    if not name or not code:
        flash("Le nom et le code de la section sont requis.", "warning")
        return redirect(url_for("dir_structure"))
    if len(code) > 20:
        flash("Le code de section ne peut pas dépasser 20 caractères.", "warning")
        return redirect(url_for("dir_structure"))
    duplicate = Section.query.filter(Section.id != section.id).filter(
        (Section.name == name) | (Section.code == code)
    ).first()
    if duplicate:
        flash("Une autre section utilise déjà ce nom ou ce code.", "warning")
        return redirect(url_for("dir_structure"))
    section.name = name
    section.code = code
    db.session.commit()
    flash("Section modifiée.", "success")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure/filiere/nouvelle", methods=["POST"])
@roles_required("directeur")
def dir_department_new():
    section_id = request.form.get("section_id", type=int)
    name = request.form.get("name", "").strip()
    code = request.form.get("code", "").strip().upper()
    if name and code and section_id:
        db.session.add(Department(name=name, code=code, section_id=section_id))
        db.session.commit()
        flash("Filière ajoutée.", "success")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure/filiere/<int:dept_id>/modifier", methods=["POST"])
@roles_required("directeur")
def dir_department_edit(dept_id):
    dept = Department.query.get_or_404(dept_id)
    dept.name = request.form.get("name", dept.name).strip()
    db.session.commit()
    flash("Filière modifiée.", "success")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure/filiere/<int:dept_id>/supprimer")
@roles_required("directeur")
def dir_department_delete(dept_id):
    dept = Department.query.get_or_404(dept_id)
    if dept.classes:
        flash(f"Impossible de supprimer « {dept.name} » : {len(dept.classes)} classe(s) y sont rattachées. Supprimez d'abord les classes.", "danger")
        return redirect(url_for("dir_structure"))
    db.session.delete(dept)
    db.session.commit()
    flash("Filière supprimée.", "info")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure/classe/nouvelle", methods=["POST"])
@roles_required("directeur")
def dir_class_new():
    dept_id = request.form.get("department_id", type=int)
    level = request.form.get("level", "").strip()
    specialty = request.form.get("specialty", "").strip()
    class_code = request.form.get("code", "").strip().upper()
    dept = Department.query.get_or_404(dept_id)
    code = specialty or dept.code
    name = f"{CLASS_LEVEL_LABELS.get(level, level)} {code}"
    if not class_code:
        flash("Le code de la classe est requis.", "warning")
    elif SchoolClass.query.filter_by(code=class_code).first():
        flash("Ce code de classe est déjà utilisé.", "warning")
    elif not SchoolClass.query.filter_by(name=name, department_id=dept_id).first():
        db.session.add(SchoolClass(name=name, code=class_code, level=level, specialty=specialty or None, department_id=dept_id,
                                    capacity=request.form.get("capacity", 48, type=int)))
        db.session.commit()
        flash(f"Classe {name} créée.", "success")
    else:
        flash("Cette classe existe déjà.", "warning")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure/classe/<int:class_id>/modifier", methods=["POST"])
@roles_required("directeur")
def dir_class_edit(class_id):
    cls = SchoolClass.query.get_or_404(class_id)
    new_level = request.form.get("level", "").strip() or cls.level
    new_specialty = request.form.get("specialty", "").strip()
    new_code = request.form.get("code", "").strip().upper()
    code = new_specialty or cls.department.code
    new_name = f"{CLASS_LEVEL_LABELS.get(new_level, new_level)} {code}"
    if not new_code:
        flash("Le code de la classe est requis.", "warning")
        return redirect(url_for("dir_structure"))
    if SchoolClass.query.filter(SchoolClass.code == new_code, SchoolClass.id != cls.id).first():
        flash("Ce code de classe est déjà utilisé.", "danger")
        return redirect(url_for("dir_structure"))
    if new_name != cls.name and SchoolClass.query.filter(
        SchoolClass.name == new_name,
        SchoolClass.department_id == cls.department_id,
        SchoolClass.id != cls.id,
    ).first():
        flash(f"Une classe « {new_name} » existe déjà dans cette filière.", "danger")
        return redirect(url_for("dir_structure"))
    cls.level = new_level
    cls.specialty = new_specialty or None
    cls.name = new_name
    cls.code = new_code
    cls.capacity = request.form.get("capacity", cls.capacity, type=int)
    db.session.commit()
    flash(f"Classe {cls.name} modifiée.", "success")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure/classe/<int:class_id>/supprimer")
@roles_required("directeur")
def dir_class_delete(class_id):
    from sqlalchemy.exc import IntegrityError
    from models import PlannedAssessment, TeacherIndicator, CustomIndicatorValue
    cls = SchoolClass.query.get_or_404(class_id)
    if cls.students:
        flash(f"Impossible de supprimer « {cls.name} » : {len(cls.students)} élève(s) y sont inscrit(s).", "danger")
        return redirect(url_for("dir_structure"))
    name = cls.name
    course_ids = [course.id for course in cls.courses]
    if course_ids:
        # Les données de planification sont liées à un cours et ne peuvent être
        # conservées sans classe. Les supprimer avant la cascade évite une
        # violation de contrainte MySQL et une erreur 500.
        PlannedAssessment.query.filter(PlannedAssessment.course_id.in_(course_ids)).delete(synchronize_session=False)
        TeacherIndicator.query.filter(TeacherIndicator.course_id.in_(course_ids)).delete(synchronize_session=False)
        CustomIndicatorValue.query.filter(CustomIndicatorValue.course_id.in_(course_ids)).delete(synchronize_session=False)
    try:
        db.session.delete(cls)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash(f"La classe « {name} » possède encore des données liées et ne peut pas être supprimée. Supprimez d’abord ses éléments associés.", "danger")
        return redirect(url_for("dir_structure"))
    flash(f"Classe {name} supprimée.", "info")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure/matiere/nouvelle", methods=["POST"])
@roles_required("censeur")
def dir_subject_new():
    from utils import user_scoped_department_ids
    name = request.form.get("name", "").strip()
    coef = request.form.get("coefficient", 1, type=int)
    category = request.form.get("category", "Enseignements Généraux")
    user = User.query.get(session["user_id"])
    target_scope = request.form.get("target_scope", "")
    scope_type, _, scope_id = target_scope.partition(":")
    if scope_type == "classe" and scope_id.isdigit():
        target_class = SchoolClass.query.get_or_404(int(scope_id))
        dept = target_class.department
    else:
        flash("Choisissez une classe avant d’ajouter la matière.", "warning")
        return redirect(url_for("dir_structure"))
    if not name:
        flash("Le nom de la matière est requis.", "warning")
        return redirect(url_for("dir_structure"))

    scoped_dept_ids = user_scoped_department_ids(user)
    if scoped_dept_ids is not None:
        # Censeur STT ou Industriel : uniquement les départements de sa section
        if dept.id not in scoped_dept_ids:
            flash("Vous ne pouvez ajouter une matière que dans les filières de votre section.", "danger")
            return redirect(url_for("dir_structure"))
    else:
        # Censeur Enseignements Généraux (portée transversale) : uniquement des matières générales, dans n'importe quelle filière
        category = "Enseignements Généraux"

    session["last_subject_class_id"] = target_class.id
    db.session.add(Subject(name=name, coefficient=coef, category=category, department_id=dept.id,
                           class_id=target_class.id))
    db.session.commit()
    flash(f"Matière ajoutée à la classe {target_class.name}.", "success")
    return redirect(url_for("dir_structure"))


def _subject_in_scope(subject, user):
    from utils import user_scoped_department_ids
    if user.role == "directeur":
        return True
    scoped_ids = user_scoped_department_ids(user)
    if scoped_ids is not None:
        return subject.department_id in scoped_ids
    if user.role == "censeur":
        # portée transversale réservée au Censeur Enseignements Généraux, mais uniquement pour les matières générales
        return subject.category == "Enseignements Généraux"
    return False  # Censeur CRM = consultation seule désormais


@app.route("/directeur/structure/matiere/<int:subject_id>/modifier", methods=["POST"])
@roles_required("directeur", "censeur", "censeur_crm")
def dir_subject_edit(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    user = User.query.get(session["user_id"])
    if not _subject_in_scope(subject, user):
        abort(403)
    subject.name = request.form.get("name", subject.name).strip()
    subject.coefficient = request.form.get("coefficient", subject.coefficient, type=int)
    subject.category = request.form.get("category", subject.category)
    db.session.commit()
    flash("Matière modifiée.", "success")
    return redirect(url_for("dir_structure"))


@app.route("/directeur/structure/matiere/<int:subject_id>/supprimer")
@roles_required("directeur", "censeur", "censeur_crm")
def dir_subject_delete(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    user = User.query.get(session["user_id"])
    if not _subject_in_scope(subject, user):
        abort(403)
    name = subject.name
    db.session.delete(subject)
    db.session.commit()
    flash(f"Matière « {name} » supprimée.", "info")
    return redirect(url_for("dir_structure"))


# ------------------------------------------------------ salles / équipements ---
def _room_in_scope(room, user):
    """Vérifie qu'une salle relève bien du périmètre du Chef des Travaux (sa section) ou du Chef de Centre CRM (CRM uniquement)."""
    if user.role == "chef_crm":
        return room.department_id is None
    if user.role == "chef_travaux":
        scoped_ids = user_scoped_department_ids(user)
        if scoped_ids is None:  # pas de section assignée : accès complet (compatibilité)
            return True
        return room.department_id in scoped_ids
    return True


def _notify_equipment_maintenance(equipment):
    """Alerte interne les responsables concernés lorsqu'un équipement devient indisponible."""
    if equipment.status not in ("En panne", "En maintenance"):
        return
    recipients = User.query.filter(User.active == True, User.role.in_(("directeur", "chef_travaux", "chef_crm"))).all()  # noqa: E712
    for recipient in recipients:
        if _room_in_scope(equipment.room, recipient):
            notify(recipient.id, f"Maintenance équipement : {equipment.name} — {equipment.status} ({equipment.room.name}).",
                   link=url_for("rooms_list"))


@app.route("/salles")
@roles_required("directeur", "censeur", "enseignant", "chef_travaux", "chef_crm")
def rooms_list():
    from utils import user_scoped_department_ids
    user = User.query.get(session["user_id"])
    rooms = Room.query.order_by(Room.type, Room.name).all()
    maintenance = MaintenanceRequest.query.order_by(MaintenanceRequest.date.desc()).limit(15).all()
    reservations = Reservation.query.order_by(Reservation.date.desc()).limit(15).all()
    scoped_dept_ids = user_scoped_department_ids(user) if user.role == "chef_travaux" else None
    scoped_departments = Department.query.filter(Department.id.in_(scoped_dept_ids)).all() if scoped_dept_ids else Department.query.all()
    return render_template("rooms_list.html", rooms=rooms, maintenance=maintenance, reservations=reservations,
                            room_in_scope=lambda r: _room_in_scope(r, user), scoped_departments=scoped_departments)


@app.route("/salles/occupation")
@roles_required("directeur", "censeur", "enseignant", "chef_travaux", "chef_crm")
def rooms_occupation():
    """Calendrier hebdomadaire combinant créneaux réguliers et réservations ponctuelles."""
    raw_week = request.args.get("semaine", "")
    try:
        anchor = date.fromisoformat(raw_week) if raw_week else date.today()
    except ValueError:
        anchor = date.today()
        flash("La semaine demandée est invalide ; la semaine en cours est affichée.", "warning")
    week_start = anchor - timedelta(days=anchor.weekday())
    week_days = [week_start + timedelta(days=offset) for offset in range(5)]
    day_names = ("Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi")
    rooms = Room.query.order_by(Room.name).all()
    occupancy = {room.id: {day.isoformat(): [] for day in week_days} for room in rooms}
    occupancy_minutes = {room.id: 0 for room in rooms}

    def slot_duration(start, end):
        try:
            sh, sm = map(int, start.split(":"))
            eh, em = map(int, end.split(":"))
            return max(0, (eh * 60 + em) - (sh * 60 + sm))
        except (TypeError, ValueError):
            return 0
    day_by_name = dict(zip(day_names, week_days))
    for entry in ScheduleEntry.query.filter(ScheduleEntry.day.in_(day_names)).all():
        slot_day = day_by_name.get(entry.day)
        if slot_day and entry.room_id in occupancy:
            occupancy[entry.room_id][slot_day.isoformat()].append({
                "kind": "Cours", "start": entry.start_time, "end": entry.end_time,
                "label": f"{entry.course.school_class.name} · {entry.course.subject.name}",
            })
            occupancy_minutes[entry.room_id] += slot_duration(entry.start_time, entry.end_time)
    reservations = Reservation.query.filter(Reservation.date >= week_days[0], Reservation.date <= week_days[-1]).all()
    for reservation in reservations:
        key = reservation.date.isoformat()
        if reservation.room_id in occupancy and key in occupancy[reservation.room_id]:
            occupancy[reservation.room_id][key].append({
                "kind": "Réservation", "start": reservation.start_time, "end": reservation.end_time,
                "label": reservation.purpose,
            })
            occupancy_minutes[reservation.room_id] += slot_duration(reservation.start_time, reservation.end_time)
    for by_day in occupancy.values():
        for slots in by_day.values():
            slots.sort(key=lambda item: item["start"])
    weekly_capacity_minutes = 5 * 10 * 60  # 5 jours, plage indicative 07:30–17:30
    room_utilisation = [{"room": room, "minutes": occupancy_minutes[room.id],
                         "rate": min(100, round(occupancy_minutes[room.id] * 100 / weekly_capacity_minutes))}
                        for room in rooms]
    return render_template("rooms_occupation.html", rooms=rooms, occupancy=occupancy, room_utilisation=room_utilisation,
                           week_days=week_days, week_start=week_start, previous_week=week_start - timedelta(days=7),
                           next_week=week_start + timedelta(days=7))


@app.route("/salles/nouvelle", methods=["POST"])
@roles_required("directeur", "chef_travaux", "chef_crm")
def room_new():
    user = User.query.get(session["user_id"])
    name = request.form.get("name", "").strip()
    room_type = request.form.get("type", "Salle").strip()
    capacity = request.form.get("capacity", type=int) or 40
    dept_id = request.form.get("department_id", type=int)
    if not name:
        flash("Le nom de la salle est obligatoire.", "warning")
        return redirect(url_for("rooms_list"))
    if room_type not in ("Salle", "Atelier", "Laboratoire"):
        flash("Le type de salle est invalide.", "warning")
        return redirect(url_for("rooms_list"))
    if capacity < 1 or capacity > 2000:
        flash("La capacité doit être comprise entre 1 et 2 000 places.", "warning")
        return redirect(url_for("rooms_list"))
    if dept_id and not Department.query.get(dept_id):
        flash("La filière sélectionnée n’existe plus.", "warning")
        return redirect(url_for("rooms_list"))
    if user.role == "chef_crm":
        dept_id = None
    elif user.role == "chef_travaux" and user.section_id:
        scoped_ids = user_scoped_department_ids(user)
        if dept_id not in (scoped_ids or []):
            abort(403)
    db.session.add(Room(name=name, type=room_type, capacity=capacity,
                         location=request.form.get("location", "").strip(), department_id=dept_id))
    db.session.commit()
    flash("Salle/atelier ajouté à l'inventaire.", "success")
    return redirect(url_for("rooms_list"))


@app.route("/salles/<int:room_id>/equipement", methods=["POST"])
@roles_required("directeur", "chef_travaux", "chef_crm")
def equipment_new(room_id):
    user = User.query.get(session["user_id"])
    room = Room.query.get_or_404(room_id)
    if not _room_in_scope(room, user):
        abort(403)
    equipment = Equipment(name=request.form.get("name"), room_id=room_id,
                          status=request.form.get("status", "Opérationnel"),
                          quantity=request.form.get("quantity", 1, type=int))
    db.session.add(equipment)
    db.session.flush()
    _notify_equipment_maintenance(equipment)
    db.session.commit()
    flash("Équipement ajouté.", "success")
    return redirect(url_for("rooms_list"))


@app.route("/salles/<int:room_id>/modifier", methods=["POST"])
@roles_required("chef_travaux", "chef_crm", "directeur")
def room_edit(room_id):
    user = User.query.get(session["user_id"])
    room = Room.query.get_or_404(room_id)
    if not _room_in_scope(room, user):
        abort(403)
    room.name = request.form.get("name", room.name).strip()
    room.type = request.form.get("type", room.type)
    room.capacity = request.form.get("capacity", room.capacity, type=int)
    room.location = request.form.get("location", room.location)
    db.session.commit()
    flash("Salle modifiée.", "success")
    return redirect(url_for("rooms_list"))


@app.route("/salles/<int:room_id>/supprimer")
@roles_required("chef_travaux", "chef_crm", "directeur")
def room_delete(room_id):
    user = User.query.get(session["user_id"])
    room = Room.query.get_or_404(room_id)
    if not _room_in_scope(room, user):
        abort(403)
    name = room.name
    db.session.delete(room)
    db.session.commit()
    flash(f"Salle « {name} » supprimée.", "info")
    return redirect(url_for("rooms_list"))


@app.route("/salles/equipement/<int:equipment_id>/modifier", methods=["POST"])
@roles_required("chef_travaux", "chef_crm", "directeur")
def equipment_edit(equipment_id):
    user = User.query.get(session["user_id"])
    eq = Equipment.query.get_or_404(equipment_id)
    if not _room_in_scope(eq.room, user):
        abort(403)
    eq.name = request.form.get("name", eq.name).strip()
    previous_status = eq.status
    eq.status = request.form.get("status", eq.status)
    eq.quantity = request.form.get("quantity", eq.quantity, type=int)
    if eq.status != previous_status:
        _notify_equipment_maintenance(eq)
    db.session.commit()
    flash("Équipement modifié.", "success")
    return redirect(url_for("rooms_list"))


@app.route("/salles/equipement/<int:equipment_id>/supprimer")
@roles_required("chef_travaux", "chef_crm", "directeur")
def equipment_delete(equipment_id):
    user = User.query.get(session["user_id"])
    eq = Equipment.query.get_or_404(equipment_id)
    if not _room_in_scope(eq.room, user):
        abort(403)
    db.session.delete(eq)
    db.session.commit()
    flash("Équipement supprimé.", "info")
    return redirect(url_for("rooms_list"))


@app.route("/salles/reservation", methods=["POST"])
@roles_required("directeur", "censeur", "enseignant", "chef_travaux", "chef_crm")
def reservation_new():
    room_id = request.form.get("room_id", type=int)
    purpose = request.form.get("purpose", "").strip()
    start_time = request.form.get("start_time", "")
    end_time = request.form.get("end_time", "")
    try:
        first_date = date.fromisoformat(request.form.get("date", ""))
    except ValueError:
        flash("La date de réservation est invalide.", "danger")
        return redirect(url_for("rooms_list"))
    repeat_weeks = min(max(request.form.get("repeat_weeks", 1, type=int) or 1, 1), 52)
    room = Room.query.get(room_id)
    if not room or not purpose or not start_time or not end_time or start_time >= end_time:
        flash("Renseignez une salle, un motif et des horaires valides.", "danger")
        return redirect(url_for("rooms_list"))
    day_names = ("Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche")
    conflicts = []
    target_dates = [first_date + timedelta(weeks=index) for index in range(repeat_weeks)]
    for target_date in target_dates:
        recurring_conflicts = check_schedule_conflict(day_names[target_date.weekday()], start_time, end_time, room_id=room_id)
        if recurring_conflicts:
            conflicts.extend([f"{target_date.strftime('%d/%m/%Y')} : {item}" for item in recurring_conflicts])
        existing = Reservation.query.filter_by(room_id=room_id, date=target_date).all()
        for reservation in existing:
            if start_time < reservation.end_time and reservation.start_time < end_time:
                conflicts.append(f"{target_date.strftime('%d/%m/%Y')} : réservation existante de {reservation.start_time} à {reservation.end_time}")
    if conflicts:
        flash("Réservation non créée — conflit : " + " | ".join(dict.fromkeys(conflicts)), "danger")
        return redirect(url_for("rooms_list"))
    for target_date in target_dates:
        db.session.add(Reservation(room_id=room_id, purpose=purpose, date=target_date,
                                   start_time=start_time, end_time=end_time, requested_by_id=session["user_id"]))
    db.session.commit()
    flash(f"{repeat_weeks} réservation(s) enregistrée(s).", "success")
    return redirect(url_for("rooms_list"))


@app.route("/salles/maintenance", methods=["POST"])
@roles_required("directeur", "censeur", "enseignant", "chef_travaux", "chef_crm")
def maintenance_new():
    record = MaintenanceRequest(room_id=request.form.get("room_id", type=int),
                                description=request.form.get("description"),
                                reported_by_id=session["user_id"])
    db.session.add(record)
    room = Room.query.get(record.room_id)
    if room:
        recipients = User.query.filter(User.active == True, User.role.in_(("directeur", "chef_travaux", "chef_crm"))).all()  # noqa: E712
        for recipient in recipients:
            if _room_in_scope(room, recipient):
                notify(recipient.id, f"Maintenance à examiner : {room.name} — {record.description}", link=url_for("rooms_list"))
    db.session.commit()
    flash("Demande de maintenance signalée.", "info")
    return redirect(url_for("rooms_list"))


@app.route("/salles/maintenance/<int:req_id>/statut", methods=["POST"])
@roles_required("chef_travaux", "chef_crm")
def maintenance_status(req_id):
    user = User.query.get(session["user_id"])
    m = MaintenanceRequest.query.get_or_404(req_id)
    if m.room and not _room_in_scope(m.room, user):
        abort(403)
    m.status = request.form.get("status", m.status)
    db.session.commit()
    return redirect(url_for("rooms_list"))


# ----------------------------------------------------- validation EDT ---
@app.route("/directeur/emplois-du-temps")
@roles_required("directeur")
def dir_schedule_validate():
    schedule_view = request.args.get("view", "classe")
    if schedule_view not in {"classe", "individuel"}:
        schedule_view = "classe"
    section_id = request.args.get("section_id", type=int)
    department_id = request.args.get("department_id", type=int)
    class_id = request.args.get("class_id", type=int)
    sections = []
    departments = []
    available_classes = []
    classes = []
    if schedule_view == "classe":
        sections = Section.query.order_by(db.func.lower(Section.name)).all()
        if section_id and section_id not in {section.id for section in sections}:
            abort(404)

        departments_q = Department.query
        if section_id:
            departments_q = departments_q.filter(Department.section_id == section_id)
        departments = departments_q.order_by(db.func.lower(Department.name)).all()
        if department_id and department_id not in {department.id for department in departments}:
            abort(404)

        classes_q = SchoolClass.query.join(Department)
        if section_id:
            classes_q = classes_q.filter(Department.section_id == section_id)
        if department_id:
            classes_q = classes_q.filter(SchoolClass.department_id == department_id)
        available_classes = classes_q.order_by(db.func.lower(SchoolClass.name)).all()
        if class_id and class_id not in {school_class.id for school_class in available_classes}:
            abort(404)
        classes = [school_class for school_class in available_classes if school_class.id == class_id] if class_id else available_classes
    teacher_query = request.args.get("q", "").strip() if schedule_view == "individuel" else ""
    teachers = []
    teacher_schedule_slots = {}
    if schedule_view == "individuel":
        teachers_q = Teacher.query.join(User)
        if teacher_query:
            teachers_q = teachers_q.filter(db.func.lower(User.full_name).like(f"%{teacher_query.lower()}%"))
        teachers = teachers_q.order_by(db.func.lower(User.full_name)).all()
        if teachers:
            from utils import build_official_grid, filled_official_slots
            teacher_ids = [teacher.id for teacher in teachers]
            entries_by_teacher = {teacher_id: [] for teacher_id in teacher_ids}
            entries = ScheduleEntry.query.join(Course).filter(Course.teacher_id.in_(teacher_ids)).all()
            for entry in entries:
                entries_by_teacher[entry.course.teacher_id].append(entry)
            teacher_schedule_slots = {
                teacher_id: filled_official_slots(build_official_grid(teacher_entries))
                for teacher_id, teacher_entries in entries_by_teacher.items()
            }
    return render_template("dir_schedule_validate.html", schedule_view=schedule_view, classes=classes, teachers=teachers,
                           sections=sections, departments=departments, available_classes=available_classes,
                           section_id=section_id, department_id=department_id, class_id=class_id,
                           teacher_query=teacher_query, teacher_schedule_slots=teacher_schedule_slots)


@app.route("/directeur/emplois-du-temps/enseignants/<int:teacher_id>")
@roles_required("directeur")
def dir_teacher_schedule_official(teacher_id):
    from utils import DAYS, OFFICIAL_PERIODS, build_official_grid, filled_official_slots
    from enseignant_routes import DAY_EN
    teacher = Teacher.query.get_or_404(teacher_id)
    entries = ScheduleEntry.query.join(Course).filter(Course.teacher_id == teacher.id).all()
    grid = build_official_grid(entries)
    planned_slots = filled_official_slots(grid)
    classes_tenues = ", ".join(sorted({course.school_class.code or course.school_class.name for course in teacher.courses}))
    return render_template("schedule_official.html", mode="individuel", teacher=teacher, grid=grid,
                           periods=OFFICIAL_PERIODS, days=DAYS[:5], day_en=DAY_EN,
                           hours_faites=planned_slots, planned_slots=planned_slots, classes_tenues=classes_tenues,
                           pdf_url=url_for("dir_teacher_schedule_official_pdf", teacher_id=teacher.id),
                           xlsx_url=url_for("dir_teacher_schedule_official_xlsx", teacher_id=teacher.id))


@app.route("/directeur/emplois-du-temps/enseignants/<int:teacher_id>/officiel.pdf")
@roles_required("directeur")
def dir_teacher_schedule_official_pdf(teacher_id):
    from flask import send_file
    from pdf_utils import render_pdf
    from utils import DAYS, OFFICIAL_PERIODS, build_official_grid, filled_official_slots
    from enseignant_routes import DAY_EN
    teacher = Teacher.query.get_or_404(teacher_id)
    entries = ScheduleEntry.query.join(Course).filter(Course.teacher_id == teacher.id).all()
    grid = build_official_grid(entries)
    planned_slots = filled_official_slots(grid)
    pdf = render_pdf("pdf/schedule_official_pdf.html", mode="individuel", teacher=teacher,
                     grid=grid, periods=OFFICIAL_PERIODS, days=DAYS[:5], day_en=DAY_EN,
                     hours_faites=planned_slots, planned_slots=planned_slots,
                     classes_tenues=", ".join(sorted({course.school_class.code or course.school_class.name for course in teacher.courses})))
    if not pdf:
        abort(500)
    filename = f"Emploi_du_temps_{teacher.user.full_name}.pdf".replace(" ", "_")
    return send_file(pdf, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/directeur/emplois-du-temps/enseignants/<int:teacher_id>/officiel.xlsx")
@roles_required("directeur")
def dir_teacher_schedule_official_xlsx(teacher_id):
    from flask import send_file
    from excel_utils import teacher_schedule_workbook
    from utils import DAYS, OFFICIAL_PERIODS, build_official_grid, filled_official_slots
    teacher = Teacher.query.get_or_404(teacher_id)
    entries = ScheduleEntry.query.join(Course).filter(Course.teacher_id == teacher.id).all()
    grid_raw = {day: [] for day in DAYS[:5]}
    for entry in entries:
        if entry.day in grid_raw:
            grid_raw[entry.day].append(entry)
    planned_slots = filled_official_slots(build_official_grid(entries))
    workbook = teacher_schedule_workbook(teacher, grid_raw, DAYS[:5], OFFICIAL_PERIODS, planned_slots=planned_slots)
    filename = f"Emploi_du_temps_{teacher.user.full_name}.xlsx".replace(" ", "_")
    return send_file(workbook, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


# ----------------------------------------------------- calendrier scolaire ---
@app.route("/directeur/calendrier")
@roles_required("directeur")
def dir_calendar():
    from models import SchoolCalendarEvent
    events = SchoolCalendarEvent.query.order_by(SchoolCalendarEvent.position).all()
    return render_template("dir_calendar.html", events=events)


@app.route("/directeur/calendrier/nouveau", methods=["POST"])
@roles_required("directeur")
def dir_calendar_new():
    from models import SchoolCalendarEvent
    label = request.form.get("label", "").strip()
    date_text = request.form.get("date_text", "").strip()
    if label and date_text:
        max_pos = db.session.query(db.func.max(SchoolCalendarEvent.position)).scalar() or 0
        db.session.add(SchoolCalendarEvent(label=label, date_text=date_text, position=max_pos + 1))
        db.session.commit()
        flash("Événement ajouté au calendrier scolaire.", "success")
    return redirect(url_for("dir_calendar"))


@app.route("/directeur/calendrier/<int:event_id>/modifier", methods=["POST"])
@roles_required("directeur")
def dir_calendar_edit(event_id):
    from models import SchoolCalendarEvent
    e = SchoolCalendarEvent.query.get_or_404(event_id)
    e.label = request.form.get("label", e.label).strip()
    e.date_text = request.form.get("date_text", e.date_text).strip()
    db.session.commit()
    flash("Événement modifié.", "success")
    return redirect(url_for("dir_calendar"))


@app.route("/directeur/calendrier/<int:event_id>/supprimer")
@roles_required("directeur")
def dir_calendar_delete(event_id):
    from models import SchoolCalendarEvent
    e = SchoolCalendarEvent.query.get_or_404(event_id)
    db.session.delete(e)
    db.session.commit()
    flash("Événement supprimé.", "info")
    return redirect(url_for("dir_calendar"))


# ----------------------------------------------------- paramètres généraux ---
@app.route("/directeur/parametres", methods=["GET", "POST"])
@roles_required("directeur")
def dir_settings():
    from models import AppSetting
    setting = AppSetting.query.first()
    if not setting:
        setting = AppSetting(current_school_year="2025-2026")
        db.session.add(setting)
        db.session.commit()
    if request.method == "POST":
        new_year = request.form.get("school_year", "").strip()
        if new_year:
            setting.current_school_year = new_year
            db.session.commit()
            flash(f"Année scolaire mise à jour : {new_year}. Elle s'applique désormais partout.", "success")
        return redirect(url_for("dir_settings"))
    return render_template("dir_settings.html", setting=setting)
