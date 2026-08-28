import os
import subprocess
from io import BytesIO
from datetime import date

from openpyxl import load_workbook

os.environ.setdefault("LTT_ENV", "development")
os.environ.setdefault("LTT_INITIAL_ADMIN_PASSWORD", "FoundateurTest#2026")

from app import app
from models import Course, Department, Grade, SchoolClass, Section, Student, Subject, Teacher, User, db
from utils import bulletin_data


def add_user(username, role):
    user = User(username=username, full_name=username, role=role, active=True)
    user.set_password("Test#2026")
    db.session.add(user)
    db.session.flush()
    return user


def add_term_grades(student, courses, value):
    for course in courses:
        db.session.add(Grade(value=value, max_value=20, student_id=student.id, course_id=course.id,
                             term="Trimestre 1", type="Évaluation", sequence=1))
        db.session.add(Grade(value=value, max_value=20, student_id=student.id, course_id=course.id,
                             term="Trimestre 1", type="Évaluation", sequence=2))


with app.app_context():
    db.drop_all()
    db.create_all()
    censeur = add_user("censeur.honneur", "censeur")
    teacher_user = add_user("enseignant.honneur", "enseignant")
    teacher_two_user = add_user("enseignant.honneur2", "enseignant")
    section = Section(name="Section Industrielle", code="IND")
    db.session.add(section)
    db.session.flush()
    department = Department(name="Génie électrique", code="GEL", section_id=section.id)
    db.session.add(department)
    db.session.flush()
    school_class = SchoolClass(name="1ère GEL", level="1ère", department_id=department.id)
    other_class = SchoolClass(name="2nde GEL", level="2nde", department_id=department.id)
    subjects = [Subject(name="Électrotechnique", coefficient=3, department_id=department.id),
                Subject(name="Mathématiques", coefficient=2, department_id=department.id)]
    teachers = [Teacher(user_id=teacher_user.id, department_id=department.id),
                Teacher(user_id=teacher_two_user.id, department_id=department.id)]
    db.session.add_all([school_class, other_class, *subjects, *teachers])
    db.session.flush()
    courses = [Course(class_id=school_class.id, subject_id=subjects[0].id, teacher_id=teachers[0].id),
               Course(class_id=school_class.id, subject_id=subjects[1].id, teacher_id=teachers[1].id)]
    other_courses = [Course(class_id=other_class.id, subject_id=subjects[0].id, teacher_id=teachers[0].id),
                     Course(class_id=other_class.id, subject_id=subjects[1].id, teacher_id=teachers[1].id)]
    db.session.add_all([*courses, *other_courses])
    db.session.flush()

    students = []
    for username, matricule, first_name, last_name, dob, birth_place in [
        ("honor.lower", "HON-012", "Aline", "Douze", date(2009, 1, 12), "Tibati"),
        ("honor.upper", "HON-020", "Bruno", "Vingt", date(2008, 9, 20), "Ngaoundéré"),
        ("honor.low", "HON-011", "Claude", "Onze", date(2009, 3, 8), "Tibati"),
        ("honor.pending", "HON-INC", "Diane", "Incomplète", date(2009, 7, 5), "Meiganga"),
    ]:
        student_user = add_user(username, "eleve")
        student = Student(user_id=student_user.id, class_id=school_class.id, matricule=matricule,
                          first_name=first_name, last_name=last_name, dob=dob, birth_place=birth_place,
                          sex="F" if first_name in {"Aline", "Diane"} else "M")
        db.session.add(student)
        students.append(student)
    db.session.flush()
    add_term_grades(students[0], courses, 12)
    add_term_grades(students[1], courses, 20)
    add_term_grades(students[2], courses, 11)
    add_term_grades(students[3], courses[:1], 15)
    other_user = add_user("honor.other", "eleve")
    other_student = Student(user_id=other_user.id, class_id=other_class.id, matricule="HON-OTH", first_name="Éric",
                            last_name="Autre", dob=date(2009, 5, 2), birth_place="Banyo", sex="M")
    db.session.add(other_student)
    db.session.flush()
    add_term_grades(other_student, other_courses, 15)
    db.session.commit()

    assert bulletin_data(students[0], "Trimestre 1")["overall_avg"] == 12
    assert bulletin_data(students[1], "Trimestre 1")["overall_avg"] == 20
    assert bulletin_data(students[2], "Trimestre 1")["overall_avg"] == 11

    with app.test_client() as client:
        client.post("/login", data={"username": "censeur.honneur", "password": "Test#2026"})
        structure = client.get("/directeur/structure")
        assert structure.status_code == 200 and "Professeur principal".encode() in structure.data
        assign = client.post(f"/directeur/structure/classe/{school_class.id}/professeur-principal",
                             data={"teacher_id": teachers[0].id})
        assert assign.status_code in (302, 303)
        assert db.session.get(SchoolClass, school_class.id).homeroom_teacher_id == teachers[0].id
        listing = client.get(f"/censeur/bulletins?class_id={school_class.id}&term=Trimestre+1")
        assert listing.status_code == 200 and "Tableaux d’honneur (12".encode() not in listing.data
        custom_text = "pour son mérite exceptionnel et son engagement exemplaire."
        preview = client.get(f"/censeur/bulletins/classe/{school_class.id}/tableaux-honneur/apercu?term=Trimestre+1&congratulations={custom_text}")
        assert preview.status_code == 200 and custom_text.encode() in preview.data
        honor_pdf = client.get(f"/censeur/bulletins/classe/{school_class.id}/tableaux-honneur.pdf?term=Trimestre+1&congratulations={custom_text}")
        assert honor_pdf.status_code == 200 and honor_pdf.data[:4] == b"%PDF"
        assert b"Tableaux_honneur" in honor_pdf.headers["Content-Disposition"].encode()
        with open("/tmp/ltt-honor-roll.pdf", "wb") as output:
            output.write(honor_pdf.data)
        page_info = subprocess.run(["pdfinfo", "/tmp/ltt-honor-roll.pdf"], check=True, capture_output=True, text=True)
        assert "Pages:           3" in page_info.stdout
        register = client.get("/censeur/tableaux-honneur/registre?term=Trimestre+1")
        assert register.status_code == 200
        assert "Nom et prénom".encode() in register.data
        assert "Aline Douze".encode() in register.data
        assert "Bruno Vingt".encode() in register.data
        assert "Éric".encode() in register.data and "Autre".encode() in register.data
        assert "Claude".encode() not in register.data and "Diane".encode() in register.data
        filtered_register = client.get(f"/censeur/tableaux-honneur/registre?term=Trimestre+1&class_id={school_class.id}")
        assert filtered_register.status_code == 200 and "Aline".encode() in filtered_register.data
        assert "Éric".encode() not in filtered_register.data and "Diane".encode() in filtered_register.data
        register_export = client.get(f"/censeur/tableaux-honneur/registre/export.xlsx?term=Trimestre+1&class_id={school_class.id}")
        assert register_export.status_code == 200 and register_export.data[:2] == b"PK"
        workbook = load_workbook(BytesIO(register_export.data), data_only=True)
        headers = [workbook.active.cell(row=4, column=index).value for index in range(1, 10)]
        assert headers == ["N°", "Nom et prénom", "Date de naissance", "Lieu de naissance", "Matricule", "Classe", "Moyenne /20", "Rang", "Effectif"]
        values = " ".join(str(cell.value) for row in workbook.active.iter_rows() for cell in row if cell.value is not None)
        assert "Aline Douze" in values and "12/01/2009" in values and "Tibati" in values
        assert "Bruno Vingt" in values and "20/09/2008" in values and "Ngaoundéré" in values
        assert "Claude Onze" not in values and "Éric Autre" not in values and "Diane" in values
        register_pdf = client.get(f"/censeur/tableaux-honneur/registre/export.pdf?term=Trimestre+1&class_id={school_class.id}")
        assert register_pdf.status_code == 200 and register_pdf.data[:4] == b"%PDF"
        with open("/tmp/ltt-honor-register.pdf", "wb") as output:
            output.write(register_pdf.data)
        bulletin_preview = client.get(f"/eleves/{students[0].id}/bulletin?term=Trimestre+1")
        certificate_preview = client.get(f"/censeur/tableaux-honneur/eleve/{students[0].id}/apercu?term=Trimestre+1")
        assert bulletin_preview.status_code == 200 and certificate_preview.status_code == 200

print("HONOR_ROLL_FEATURE_TEST_OK")
