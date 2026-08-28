from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "0B2545"
GOLD = "C9A227"
CREAM = "F8F5EC"
THIN = Side(style="thin", color="D9D0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def import_report_workbook(report):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rapport d’import"
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 88
    sheet.merge_cells("A1:B1")
    title = sheet["A1"]
    title.value = "LYCÉE TECHNIQUE DE TIBATI — RAPPORT D’IMPORT"
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="center")
    sheet["A3"] = "Type d’import"; sheet["B3"] = report.get("kind", "—").capitalize()
    sheet["A4"] = "Lignes importées"; sheet["B4"] = report.get("created", 0)
    sheet["A5"] = "Lignes ignorées"; sheet["B5"] = report.get("skipped", 0)
    for row in range(3, 6):
        sheet.cell(row=row, column=1).font = Font(bold=True, color=NAVY)
    sheet["A7"] = "Statut"; sheet["B7"] = "Détail"
    for cell in sheet[7]:
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=CREAM)
        cell.border = BORDER
    errors = report.get("errors", [])
    if errors:
        for row_number, error in enumerate(errors, start=8):
            sheet.cell(row=row_number, column=1, value="À corriger").border = BORDER
            sheet.cell(row=row_number, column=2, value=error).border = BORDER
    else:
        sheet["A8"] = "Conforme"; sheet["B8"] = "Aucune ligne à corriger."
        sheet["A8"].border = BORDER; sheet["B8"].border = BORDER
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
